import json
import random
import string
import copy
from zoneinfo import ZoneInfo
from django.apps import apps
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden, HttpResponseNotAllowed, HttpResponse, Http404
from django.urls import reverse_lazy, reverse
from django.views.decorators.http import require_POST, require_GET
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.core.serializers.json import DjangoJSONEncoder
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from django.views.decorators.http import require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.views.generic import TemplateView
from django.utils.decorators import method_decorator
from django.template.loader import render_to_string
import requests
from io import TextIOWrapper
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False
from datetime import date, datetime, timedelta
import calendar

# Keep a module-level reference to datetime for attributes like strptime
import datetime as datetime_module
# Backwards-compatible alias used across legacy helpers
dt = datetime_module

from estateApp.models import *
from .models import *
from .utils import send_email_message, send_sms_message, send_inapp_message
import csv

import logging
logger = logging.getLogger(__name__)

import os
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.utils.text import get_valid_filename

from decimal import Decimal

from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum, Count, IntegerField, DecimalField, BooleanField, F, Value, Case, When, ExpressionWrapper, Subquery, OuterRef, Max
from django.db.models.functions import Coalesce
from django.views.decorators.csrf import csrf_protect
from rest_framework.authtoken.models import Token

try:
    import bleach
    BLEACH_AVAILABLE = True
except ImportError:
    BLEACH_AVAILABLE = False


User = get_user_model()
User = apps.get_model('estateApp', 'CustomUser')


SUPPORT_ROLES = ('admin', 'support')


LAGOS_TZ = ZoneInfo('Africa/Lagos')


def lagos_now():
    """Return current timezone-aware datetime in Africa/Lagos."""
    return timezone.now().astimezone(LAGOS_TZ)


def lagos_today():
    """Return today's date in Africa/Lagos."""
    return lagos_now().date()


def support_required(func=None):
    """Decorator restricting access to support/admin/staff users."""
    def _check(user):
        if not user or not user.is_authenticated:
            return False
        role = getattr(user, 'role', None)
        return role in SUPPORT_ROLES or user.is_staff or user.is_superuser

    if func:
        return user_passes_test(_check)(func)
    return user_passes_test(_check)


# @login_required
# def support_dashboard(request):
#     return render(request, 'customer_relation/support_dashboard.html',)

@login_required
def contentManagementPage(request):
    return render(request, 'adminSupport/content_management/management_content.html',)


@login_required
def manageContentPage(request):
    return render(request, 'adminSupport/content_management/manage_content_page.html',)

@login_required
def messages(request):
    return render(request, 'adminSupport/customer_relation/staff_dashboard.html',)

@login_required
def newsletter(request):
    ctx = {'home_url': '/', 'home_url': '/'}
    return render(request, 'adminSupport/customer_relation/newsletter.html',ctx)


@login_required
def autobirthday(request):
    ctx = {'home_url': '/'}
    return render(request, 'adminSupport/customer_relation/auto_birthday_templates.html', ctx)

@login_required
def autoSpecialDay(request):
    ctx = {'home_url': '/'}
    return render(request, 'adminSupport/customer_relation/auto_special_day_template.html', ctx)

def _build_initials(name: str) -> str:
    if not name:
        return "--"
    parts = name.strip().split()
    if len(parts) >= 2:
        return (parts[0][0] + parts[1][0]).upper()
    return name[:2].upper()


def _build_last_message_preview(message):
    if not message:
        return "No messages yet."

    if message.content:
        preview = message.content.strip()
        return preview if len(preview) <= 80 else preview[:77] + "..."

    if message.file:
        file_name = message.file.name.rsplit('/', 1)[-1]
        return f"Attachment: {file_name}"

    return "Sent a message"


def _authenticate_support_request(request):
    user = getattr(request, 'user', None)
    if user and getattr(user, 'is_authenticated', False) and user.role in SUPPORT_ROLES:
        return user

    auth_header = request.META.get('HTTP_AUTHORIZATION') or request.headers.get('Authorization')
    if auth_header and auth_header.lower().startswith('token '):
        token_key = auth_header.split(' ', 1)[1].strip()
        try:
            token = Token.objects.select_related('user').get(key=token_key)
        except Token.DoesNotExist:
            return None

        token_user = token.user
        if token_user.role in SUPPORT_ROLES:
            return token_user

    return None


def _collect_chat_rows(role: str):
    participants = (
        CustomUser.objects
        .filter(role=role, sent_messages__recipient__role__in=SUPPORT_ROLES)
        .distinct()
        .annotate(last_message_at=Max('sent_messages__date_sent'))
    )

    rows = []
    for user in participants:
        conversation_qs = Message.objects.filter(
            Q(sender=user, recipient__role__in=SUPPORT_ROLES) |
            Q(sender__role__in=SUPPORT_ROLES, recipient=user)
        ).order_by('-date_sent')

        last_message = conversation_qs.first()
        last_ts = last_message.date_sent if last_message else user.last_message_at
        if last_ts is None:
            last_ts = timezone.now()

        full_name = user.get_full_name() or user.username or "Unknown"
        last_seen = user.last_login or user.last_message_at or user.date_registered
        rows.append({
            'id': user.id,
            'name': full_name,
            'first_name': user.first_name or '',
            'last_name': user.last_name or '',
            'initials': _build_initials(full_name),
            'avatar_url': user.profile_image.url if getattr(user, 'profile_image', None) else '',
            'last_seen': last_seen,
            'unread_count': Message.objects.filter(
                sender=user,
                recipient__role__in=SUPPORT_ROLES,
                is_read=False,
            ).count(),
            'last_message': _build_last_message_preview(last_message),
            'last_message_ts': last_ts,
            'last_message_timestamp_iso': last_ts.isoformat() if last_ts else '',
            'url': reverse('adminSupport:chat_conversation', args=[role, user.id]),
            'role': role,
        })
    return rows


def _get_participant_or_404(role: str, user_id: int):
    normalized = (role or '').lower()
    if normalized not in ('client', 'marketer'):
        raise Http404("Invalid conversation type")
    return get_object_or_404(CustomUser, id=user_id, role=normalized)


def _conversation_queryset(participant):
    return Message.objects.filter(
        Q(sender=participant, recipient__role__in=SUPPORT_ROLES) |
        Q(sender__role__in=SUPPORT_ROLES, recipient=participant)
    ).order_by('date_sent')


def _mark_conversation_read(participant):
    Message.objects.filter(
        sender=participant,
        recipient__role__in=SUPPORT_ROLES,
        is_read=False
    ).update(is_read=True, status='read')


def _render_messages_html(request, messages):
    fragments = []
    for msg in messages:
        fragments.append(render_to_string('adminSupport/chat_message.html', {'msg': msg, 'request': request}))
    return ''.join(fragments)


def _build_poll_response(request, conversation, last_msg_id):
    try:
        last_msg_id = int(last_msg_id or 0)
    except (TypeError, ValueError):
        last_msg_id = 0

    new_messages = conversation.filter(id__gt=last_msg_id)
    messages_html = _render_messages_html(request, new_messages)
    messages_list = [{'id': msg.id} for msg in new_messages]
    updated_statuses = [{'id': msg.id, 'status': msg.status} for msg in conversation]

    return JsonResponse({
        'messages': messages_list,
        'messages_html': messages_html,
        'updated_statuses': updated_statuses,
    })


def _handle_support_message_send(request, participant):
    message_content = request.POST.get('message_content')
    if message_content is None:
        message_content = request.POST.get('content', '')
    message_content = (message_content or '').strip()
    file_attachment = request.FILES.get('file')

    if not message_content and not file_attachment:
        return JsonResponse({'success': False, 'error': 'Please enter a message or attach a file.'}, status=400)

    new_message = Message.objects.create(
        sender=request.user,
        recipient=participant,
        message_type="enquiry",
        content=message_content,
        file=file_attachment,
        status='sent'
    )

    message_html = render_to_string('adminSupport/chat_message.html', {'msg': new_message, 'request': request})
    return JsonResponse({'success': True, 'message_html': message_html, 'message_id': new_message.id})


@login_required
@support_required
def chat_conversation(request, role, user_id):
    participant = _get_participant_or_404(role, user_id)
    _mark_conversation_read(participant)
    conversation = _conversation_queryset(participant)
    last_message = conversation.last()

    if request.method == "POST":
        return _handle_support_message_send(request, participant)

    context = {
        'client': participant,
        'messages': conversation,
        'is_marketer': participant.role == 'marketer',
        'conversation_role': participant.role,
        'last_message_id': last_message.id if last_message else 0,
    }
    return render(request, 'adminSupport/chat_interface.html', context)


@login_required
@support_required
@require_GET
def chat_poll(request, role, user_id):
    participant = _get_participant_or_404(role, user_id)
    _mark_conversation_read(participant)
    conversation = _conversation_queryset(participant)
    last_msg = request.GET.get('last_msg')
    return _build_poll_response(request, conversation, last_msg)


@login_required
@support_required
@require_POST
def chat_send_message(request, role, user_id):
    participant = _get_participant_or_404(role, user_id)
    return _handle_support_message_send(request, participant)


def _search_users_by_role(role, query):
    query = (query or '').strip()
    if len(query) < 2:
        return CustomUser.objects.none()

    filters = Q(full_name__icontains=query) | Q(email__icontains=query) | Q(phone__icontains=query)
    return (
        CustomUser.objects
        .filter(role=role)
        .filter(filters)
        .only('id', 'full_name', 'email')
        .order_by('full_name')[:10]
    )


@require_GET
def chat_search_clients(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'clients': []})

    clients = _search_users_by_role('client', query)
    data = []
    for client in clients:
        full_name = client.full_name or client.get_full_name() or client.username or "Client"
        profile_image = getattr(client.profile_image, 'url', None) if getattr(client, 'profile_image', None) else None
        if profile_image and profile_image.startswith('/'):
            profile_image = request.build_absolute_uri(profile_image)
        data.append({
            'id': client.id,
            'full_name': full_name,
            'email': client.email or '',
            'profile_image': profile_image,
        })

    return JsonResponse({'clients': data})


@require_GET
def chat_search_marketers(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'marketers': []})

    marketers = _search_users_by_role('marketer', query)
    data = []
    for marketer in marketers:
        full_name = marketer.full_name or marketer.get_full_name() or marketer.username or "Marketer"
        profile_image = getattr(marketer.profile_image, 'url', None) if getattr(marketer, 'profile_image', None) else None
        if profile_image and profile_image.startswith('/'):
            profile_image = request.build_absolute_uri(profile_image)
        data.append({
            'id': marketer.id,
            'full_name': full_name,
            'email': marketer.email or '',
            'profile_image': profile_image,
        })

    return JsonResponse({'marketers': data})


def _serialize_support_chat_row(row):
    timestamp = row.get('last_message_ts')
    timestamp_iso_str = row.get('last_message_timestamp_iso')

    timestamp_dt = None

    if isinstance(timestamp, dt.datetime):
        timestamp_dt = timestamp
    elif isinstance(timestamp, dt.date):
        timestamp_dt = dt.datetime.combine(timestamp, dt.time.min)
    elif isinstance(timestamp_iso_str, str) and timestamp_iso_str:
        try:
            timestamp_dt = dt.datetime.fromisoformat(timestamp_iso_str)
        except ValueError:
            timestamp_dt = None

    if timestamp_dt is None:
        timestamp_dt = timezone.now()

    if timezone.is_naive(timestamp_dt):
        timestamp_dt = timezone.make_aware(timestamp_dt, timezone.get_current_timezone())

    timestamp_iso = timestamp_dt.isoformat()

    name = row.get('name') or ''
    first_name = row.get('first_name') or ''
    last_name = row.get('last_name') or ''

    if not first_name and name:
        parts = name.split(' ', 1)
        first_name = parts[0]
        if len(parts) > 1 and not last_name:
            last_name = parts[1]

    profile_image = row.get('profile_image') or row.get('profile_image_url')
    if profile_image and isinstance(profile_image, str) and profile_image.startswith('/'):
        profile_image = request.build_absolute_uri(profile_image)

    payload = {
        'id': row.get('id'),
        'first_name': first_name,
        'last_name': last_name,
        'last_message': row.get('last_message', ''),
        'unread_count': row.get('unread_count', 0),
        'timestamp': timestamp_iso,
        'profile_image': profile_image,
    }

    return payload


def chat_list_clients_api(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    rows = _collect_chat_rows('client')
    payload = [_serialize_support_chat_row(row) for row in rows]
    return JsonResponse(payload, safe=False)


def chat_list_marketers_api(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    rows = _collect_chat_rows('marketer')
    payload = [_serialize_support_chat_row(row) for row in rows]
    return JsonResponse(payload, safe=False)


@login_required
@support_required
def chat(request):
    client_rows = _collect_chat_rows('client')
    marketer_rows = _collect_chat_rows('marketer')

    client_unread_threads = sum(1 for row in client_rows if row.get('unread_count', 0))
    marketer_unread_threads = sum(1 for row in marketer_rows if row.get('unread_count', 0))
    client_unread_messages = sum(row.get('unread_count', 0) for row in client_rows)
    marketer_unread_messages = sum(row.get('unread_count', 0) for row in marketer_rows)

    context = {
        'client_rows': client_rows,
        'marketer_rows': marketer_rows,
        'total_unread_clients': client_unread_messages,
        'total_unread_marketers': marketer_unread_messages,
        'total_unread_threads': client_unread_threads + marketer_unread_threads,
        'total_unread_messages': client_unread_messages + marketer_unread_messages,
        'unread_count': client_unread_threads + marketer_unread_threads,
    }
    return render(request, 'adminSupport/chat_list.html', context)


@login_required
@support_required
def chat_list_partial(request):
    client_rows = _collect_chat_rows('client')
    marketer_rows = _collect_chat_rows('marketer')

    client_unread_threads = sum(1 for row in client_rows if row.get('unread_count', 0))
    marketer_unread_threads = sum(1 for row in marketer_rows if row.get('unread_count', 0))
    client_unread_messages = sum(row.get('unread_count', 0) for row in client_rows)
    marketer_unread_messages = sum(row.get('unread_count', 0) for row in marketer_rows)

    client_html = render_to_string(
        'adminSupport/chat_list_rows.html',
        {
            'chat_rows': client_rows,
            'empty_title': 'No Client Conversations',
            'empty_description': "When clients start conversations, they'll appear here.",
            'empty_icon': 'fa-comments',
        },
        request=request,
    )

    marketer_html = render_to_string(
        'adminSupport/chat_list_rows.html',
        {
            'chat_rows': marketer_rows,
            'empty_title': 'No Marketer Conversations',
            'empty_description': "When marketers start conversations, they'll appear here.",
            'empty_icon': 'fa-bullhorn',
        },
        request=request,
    )

    return JsonResponse({
        'clients': client_html,
        'marketers': marketer_html,
        'totals': {
            'client_messages': client_unread_messages,
            'marketer_messages': marketer_unread_messages,
            'client_threads': client_unread_threads,
            'marketer_threads': marketer_unread_threads,
        }
    })


@login_required
@support_required
@require_POST
def chat_delete_message(request):
    message_id = request.POST.get('message_id')
    if not message_id:
        return JsonResponse({'success': False, 'error': 'Message ID is required'}, status=400)

    try:
        msg = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Message not found'}, status=404)

    if msg.sender.role not in SUPPORT_ROLES and msg.recipient.role not in SUPPORT_ROLES:
        return JsonResponse({'success': False, 'error': 'You do not have permission to delete this message'}, status=403)

    msg.delete()
    return JsonResponse({'success': True, 'removed_id': int(message_id)})


# MESSAGING APP


@login_required
@support_required
def clients_directory(request):
    """Standalone Clients Directory page (moved out of dashboard tabs)."""
    q = (request.GET.get('q') or '').strip()
    qs = ClientUser.with_investment_metrics()
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q)
        )
    qs = qs.select_related('assigned_marketer').order_by('full_name')

    # optional export
    if request.GET.get('export'):
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="clients.csv"'
        writer = csv.writer(resp)
        writer.writerow(['Full Name', 'Email', 'Phone', 'Assigned Marketer', 'Date Registered'])
        for u in qs:
            writer.writerow([
                u.full_name,
                u.email,
                u.phone,
                (u.assigned_marketer.full_name if u.assigned_marketer else ''),
                u.date_registered.isoformat() if u.date_registered else ''
            ])
        return resp
    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    ctx = {
        'page_obj': page_obj,
        'q': q,
        'total': paginator.count,
    }
    return render(request, 'adminSupport/customer_relation/clients_directory.html', ctx)


@login_required
@support_required
def marketers_directory(request):
    """Standalone Marketers Directory page (moved out of dashboard tabs)."""
    q = (request.GET.get('q') or '').strip()
    current_year = timezone.now().year
    year_str = str(current_year)

    # Annual targets: prefer marketer-specific, else global (marketer=None)
    target_specific_sq = (MarketerTarget.objects
                          .filter(period_type='annual', specific_period=year_str, marketer=OuterRef('pk'))
                          .order_by('-created_at')
                          .values('target_amount')[:1])
    target_global_sq = (MarketerTarget.objects
                        .filter(period_type='annual', specific_period=year_str, marketer__isnull=True)
                        .order_by('-created_at')
                        .values('target_amount')[:1])

    qs = (MarketerUser.objects
          .annotate(
              year_sales=Coalesce(
                  Sum('transaction__total_amount', filter=Q(transaction__transaction_date__year=current_year)),
                  Value(0),
                  output_field=DecimalField(max_digits=18, decimal_places=2)
              ),
              year_deals=Coalesce(
                  Count('transaction', filter=Q(transaction__transaction_date__year=current_year), distinct=True),
                  Value(0),
                  output_field=IntegerField()
              ),
              year_target=Coalesce(
                  Subquery(target_specific_sq),
                  Subquery(target_global_sq),
                  Value(0),
                  output_field=DecimalField(max_digits=18, decimal_places=2)
              ),
          ))

    # Compute percent of target safely
    pct_expr = ExpressionWrapper(
        (F('year_sales') * Value(100.0)) / F('year_target'),
        output_field=DecimalField(max_digits=7, decimal_places=2)
    )
    qs = qs.annotate(
        year_target_pct=Case(
            When(year_target__gt=0, then=pct_expr),
            default=Value(None),
            output_field=DecimalField(max_digits=7, decimal_places=2)
        ),
        has_progress=Case(
            When(year_target__gt=0, year_sales__gt=0, then=Value(True)),
            default=Value(False),
            output_field=BooleanField()
        )
    )
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q) |
            Q(email__icontains=q) |
            Q(phone__icontains=q)
        )

    # Always order by performance
    qs = qs.order_by('-year_sales', 'full_name')

    # Determine top-3 among those who have a target and >0% progress
    eligible_ids = list(
        qs.filter(year_target__gt=0, year_target_pct__gt=0)
          .values_list('id', flat=True)[:3]
    )
    gold_ids = eligible_ids[:1]
    silver_ids = eligible_ids[1:2]
    bronze_ids = eligible_ids[2:3]

    # optional export
    if request.GET.get('export'):
        resp = HttpResponse(content_type='text/csv')
        resp['Content-Disposition'] = 'attachment; filename="marketers.csv"'
        writer = csv.writer(resp)
        writer.writerow(['Full Name', 'Email', 'Phone', 'Year', 'Deals (Year)', 'Sales (Year ₦)', 'Target (Year ₦)', 'Target %', 'Rank/Badge', 'Date Registered'])
        data = list(qs)
        for u in data:
            if u.id in gold_ids:
                badge = 'Gold'
            elif u.id in silver_ids:
                badge = 'Elite Marketer'
            elif u.id in bronze_ids:
                badge = 'Consistent Performer'
            else:
                badge = ''
            writer.writerow([
                u.full_name,
                u.email,
                u.phone,
                current_year,
                getattr(u, 'year_deals', 0),
                getattr(u, 'year_sales', 0),
                getattr(u, 'year_target', 0),
                getattr(u, 'year_target_pct', None),
                badge,
                u.date_registered.isoformat() if u.date_registered else ''
            ])
        return resp
    paginator = Paginator(qs, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    ctx = {
        'page_obj': page_obj,
        'q': q,
        'total': paginator.count,
        'current_year': current_year,
        'gold_ids': gold_ids,
        'silver_ids': silver_ids,
        'bronze_ids': bronze_ids,
    }
    return render(request, 'adminSupport/customer_relation/marketers_directory.html', ctx)


@method_decorator([login_required, support_required], name='dispatch')
class SupportMessagingDashboardView(TemplateView):
    template_name = 'adminSupport/customer_relation/support_dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['home_url'] = '/'

        # server-side fallback stats for initial render (JS will refresh)
        clients_count = User.objects.filter(role='client').count()
        marketers_count = User.objects.filter(role='marketer').count()
        ctx['stat_customers'] = clients_count + marketers_count

        # birthdays this month (guard against null date_of_birth)
        try:
            current_month = lagos_today().month
            ctx['stat_birthdays'] = User.objects.filter(date_of_birth__month=current_month).count()
        except Exception:
            ctx['stat_birthdays'] = 0

        # messages sent: outbound (not queued) + inapp published
        try:
            ctx['stat_sent'] = (
                OutboundMessage.objects.exclude(status='queued').count()
                + InAppMessage.objects.filter(is_draft=False).count()
            )
        except Exception:
            ctx['stat_sent'] = 0

        # open rate (inapp inbox entries)
        delivered = InboxEntry.objects.count()
        read = InboxEntry.objects.filter(is_read=True).count()
        if delivered:
            try:
                ctx['stat_open'] = f"{int((read / delivered) * 100)}%"
            except Exception:
                ctx['stat_open'] = "0%"
        else:
            ctx['stat_open'] = "0%"

        return ctx


def _serialize_user(u):
    # prefer explicit first_name/last_name, but fall back to full_name
    full_raw = (getattr(u, 'full_name', None) or '').strip()
    if not full_raw:
        # fallback to Django's get_full_name() or email username
        full_raw = (u.get_full_name() or (u.email.split('@')[0] if getattr(u, 'email', None) else '')).strip()

    parts = full_raw.split()
    # first name: either the explicit first_name or the first token of full_raw
    first = (u.first_name or '') or (parts[0] if parts else '')
    # last name: explicit last_name or the remainder of full_raw after the first token
    last = (u.last_name or '') or (' '.join(parts[1:]) if len(parts) > 1 else '')

    # fullName WITHOUT the first name (everything after the first token)
    full_without_first = ' '.join(parts[1:]).strip() if len(parts) > 1 else ''

    return {
        'id': u.id,
        'firstName': first,
        'lastName': last,
        'fullName': full_without_first,
        'email': getattr(u, 'email', '') or '',
        'phone': getattr(u, 'phone', '') or '',
        'role': getattr(u, 'role', '') or '',
        'birthday': u.date_of_birth.isoformat() if getattr(u, 'date_of_birth', None) else None,
    }


@require_http_methods(['GET'])
@login_required
@support_required
def api_users(request):
    """
    GET ?role=clients|marketers|all
    """
    role = request.GET.get('role', 'all')
    qs = User.objects.all()
    if role == 'clients':
        qs = qs.filter(role='client')
    elif role == 'marketers':
        qs = qs.filter(role='marketer')
    users = [_serialize_user(u) for u in qs.order_by('first_name')[:200]]
    return JsonResponse({'users': users}, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_calendar_events(request):
    """
    Return calendar events for a requested year/month.
    Query params: ?year=2025&month=9
    """
    try:
        today_lagos = lagos_today()
        year = int(request.GET.get('year', today_lagos.year))
        month = int(request.GET.get('month', today_lagos.month))
    except Exception:
        today_lagos = lagos_today()
        year = today_lagos.year
        month = today_lagos.month

    events = []

    # birthdays in requested month
    users = User.objects.exclude(date_of_birth__isnull=True).filter(date_of_birth__month=month)
    for u in users:
        try:
            dt = date(year, u.date_of_birth.month, u.date_of_birth.day)
            title = (getattr(u, 'full_name', '') or u.get_full_name() or u.email or 'User')
            events.append({'title': f"{title} ({u.role or ''})", 'start': dt.isoformat(), 'category': 'birthday'})
        except Exception:
            continue

    # outbounds scheduled in this month
    out_qs = OutboundMessage.objects.filter(scheduled_for__year=year, scheduled_for__month=month)
    for o in out_qs:
        if o.scheduled_for:
            events.append({'title': f"Outbound — {o.channel} ({o.recipients_count})", 'start': o.scheduled_for.date().isoformat(), 'category': 'outbound'})

    # optional: holidays/specials can be added server-side if you want to proxy them
    return JsonResponse({'events': events}, encoder=DjangoJSONEncoder)



@require_http_methods(['GET'])
@login_required
@support_required
def api_templates_list(request):
    q = MessageTemplate.objects.all().order_by('-created_at')[:200]
    templates = []
    for t in q:
        templates.append({
            'id': t.id,
            'name': t.name,
            'channel': t.channel,
            'audience': t.audience,
            'subject': t.subject,
            'message': t.body,
            'time': t.send_time.isoformat() if t.send_time else None,
            'createdAt': t.created_at.isoformat() if t.created_at else None,
            'isActive': t.is_active
        })
    return JsonResponse({'templates': templates}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_templates_save(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    tpl_id = payload.get('id')
    if tpl_id:
        tpl = get_object_or_404(MessageTemplate, id=tpl_id)
    else:
        tpl = MessageTemplate(created_by=request.user)

    tpl.name = payload.get('name', tpl.name or 'Template')
    tpl.channel = payload.get('channel', tpl.channel or 'inapp')
    tpl.audience = payload.get('audience', tpl.audience or 'clients')
    tpl.subject = payload.get('subject', tpl.subject)
    tpl.body = payload.get('message', tpl.body or '')

    time_str = payload.get('time')
    if time_str:
        try:
            # expect "HH:MM"
            tpl.send_time = datetime.strptime(time_str, "%H:%M").time()
        except Exception:
            pass

    tpl.is_active = payload.get('isActive', tpl.is_active)
    tpl.save()
    ActivityLog.objects.create(actor=request.user, action=f"Saved template '{tpl.name}'")
    return JsonResponse({'ok': True, 'id': tpl.id}, encoder=DjangoJSONEncoder)
    


@require_http_methods(['POST'])
@login_required
@support_required
def api_templates_delete(request, tpl_id):
    tpl = get_object_or_404(MessageTemplate, id=tpl_id)
    name = tpl.name
    tpl.delete()
    ActivityLog.objects.create(actor=request.user, action=f"Deleted template '{name}'")
    return JsonResponse({'ok': True})


@require_http_methods(['GET'])
@login_required
@support_required
def api_outbound_list(request):
    q = OutboundMessage.objects.all().order_by('-created_at')[:200]
    out = []
    for it in q:
        out.append({
            'id': str(it.id),
            'channel': it.channel,
            'audience': it.audience,
            'subject': it.subject,
            'body': it.body,
            'status': it.status,
            'createdAt': it.created_at.isoformat() if it.created_at else None,
            'scheduledFor': it.scheduled_for.isoformat() if it.scheduled_for else None,
            'toCount': it.recipients_count,
            'toSample': it.recipients_sample or [],
        })
    return JsonResponse({'outbound': out}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_outbound_create(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    channel = payload.get('channel', 'inapp')
    audience = payload.get('audience', 'clients')
    subject = payload.get('subject', '') or None
    body = payload.get('body', '') or payload.get('message', '')

    # compute recipients
    if audience == 'clients':
        recipients_qs = User.objects.filter(role='client')
    elif audience == 'marketers':
        recipients_qs = User.objects.filter(role='marketer')
    else:
        recipients_qs = User.objects.all()

    recipients_count = recipients_qs.count()
    sample = list(recipients_qs.values_list('first_name', 'last_name')[:3])
    sample_strs = [f"{a or ''} {b or ''}".strip() for a, b in sample]

    sched = payload.get('scheduled_for')
    scheduled_for_dt = None
    if sched:
        try:
            parsed = datetime.fromisoformat(sched)
            if timezone.is_naive(parsed):
                scheduled_for_dt = timezone.make_aware(parsed, timezone.get_current_timezone())
            else:
                scheduled_for_dt = parsed
        except Exception:
            scheduled_for_dt = None

    om = OutboundMessage.objects.create(
        channel=channel,
        audience=audience,
        subject=subject,
        body=body,
        created_by=request.user,
        recipients_count=recipients_count,
        recipients_sample=sample_strs,
        scheduled_for=scheduled_for_dt,
        status='queued'
    )
    ActivityLog.objects.create(actor=request.user, action=f"Queued outbound {om.channel} to {recipients_count} (id={om.id})")
    return JsonResponse({'ok': True, 'id': str(om.id)}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_outbound_delete(request, outbound_id):
    om = get_object_or_404(OutboundMessage, id=outbound_id)
    om.delete()
    ActivityLog.objects.create(actor=request.user, action=f"Deleted outbound {outbound_id}")
    return JsonResponse({'ok': True})


@require_http_methods(['GET'])
@login_required
@support_required
def api_messages_list(request):
    drafts = InAppMessage.objects.filter(is_draft=True, sender=request.user).order_by('-created_at')[:50]
    sent = InAppMessage.objects.filter(is_draft=False, sender=request.user).order_by('-created_at')[:200]

    def s(m):
        return {'id': m.id, 'subject': m.subject, 'body': m.body, 'isDraft': m.is_draft, 'createdAt': m.created_at.isoformat() if m.created_at else None}

    messages = [s(x) for x in list(drafts) + list(sent)]
    return JsonResponse({'messages': messages}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_messages_create_or_update(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    msg_id = payload.get('id')
    if msg_id:
        msg = get_object_or_404(InAppMessage, id=msg_id, sender=request.user)
    else:
        msg = InAppMessage(sender=request.user)

    msg.subject = payload.get('subject', msg.subject)
    msg.body = payload.get('body', msg.body)
    msg.is_draft = bool(payload.get('is_draft', msg.is_draft))
    msg.save()

    recipients = payload.get('recipients')
    if isinstance(recipients, str):
        if recipients == 'all':
            qs = User.objects.all()
        elif recipients == 'clients':
            qs = User.objects.filter(role='client')
        elif recipients == 'marketers':
            qs = User.objects.filter(role='marketer')
        else:
            qs = User.objects.none()
        msg.recipients.set(qs)
    elif isinstance(recipients, list):
        try:
            ids = [int(x) for x in recipients]
        except Exception:
            ids = []
        qs = User.objects.filter(id__in=ids)
        msg.recipients.set(qs)

    # make inbox entries when not draft
    if not msg.is_draft:
        entries = []
        for u in msg.recipients.all():
            entries.append(InboxEntry(user=u, message=msg))
        # ignore_conflicts available in modern Django
        InboxEntry.objects.bulk_create(entries, ignore_conflicts=True)
        ActivityLog.objects.create(actor=request.user, action=f"Sent InApp '{msg.subject or '(no subject)'}' to {msg.recipients.count()}")
    else:
        ActivityLog.objects.create(actor=request.user, action=f"Saved draft '{msg.subject or '(no subject)'}'")

    return JsonResponse({'ok': True, 'id': msg.id}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_messages_delete(request, msg_id):
    msg = get_object_or_404(InAppMessage, id=msg_id, sender=request.user)
    msg.delete()
    ActivityLog.objects.create(actor=request.user, action=f"Deleted message {msg_id}")
    return JsonResponse({'ok': True})


@require_http_methods(['GET'])
@login_required
@support_required
def api_activity(request):
    logs = ActivityLog.objects.all().order_by('-created_at')[:200]
    data = [{'at': l.created_at.isoformat() if l.created_at else None, 'text': l.action, 'actor': l.actor.get_full_name() if l.actor else None} for l in logs]
    return JsonResponse({'activity': data}, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_stats(request):
    # Total customers = clients + marketers as requested
    clients = User.objects.filter(role='client').count()
    marketers = User.objects.filter(role='marketer').count()
    total_customers = clients + marketers

    # birthdays in current month
    try:
        birthdays_month = User.objects.filter(date_of_birth__month=lagos_today().month).count()
    except Exception:
        birthdays_month = 0

    # Today's birthdays (Clients + Marketers derived from their concrete models)
    try:
        today = lagos_today()
        clients_today = ClientUser.objects.filter(
            date_of_birth__month=today.month,
            date_of_birth__day=today.day
        ).count()
        marketers_today = MarketerUser.objects.filter(
            date_of_birth__month=today.month,
            date_of_birth__day=today.day
        ).count()
        today_birthdays_count = clients_today + marketers_today
    except Exception:
        today_birthdays_count = 0

    # Upcoming birthdays (calendar week Sun–Sat) and next week
    try:
        today = lagos_today()
        client_dobs = list(ClientUser.objects.exclude(date_of_birth__isnull=True).values_list('date_of_birth', flat=True))
        marketer_dobs = list(MarketerUser.objects.exclude(date_of_birth__isnull=True).values_list('date_of_birth', flat=True))
        all_dobs = client_dobs + marketer_dobs
        # this week: Sunday..Saturday of current week
        start_week = today - timedelta(days=(today.weekday() + 1) % 7)
        end_week = start_week + timedelta(days=6)
        # next week: following Sunday..Saturday
        start_next = start_week + timedelta(days=7)
        end_next = start_next + timedelta(days=6)
        upcoming_count = 0
        next_week_count = 0
        for b in all_dobs:
            # Build window-specific dates to handle year boundaries correctly
            try:
                week_bday = datetime(year=start_week.year, month=b.month, day=b.day).date()
            except Exception:
                week_bday = None
            try:
                next_bday = datetime(year=start_next.year, month=b.month, day=b.day).date()
            except Exception:
                next_bday = None
            if week_bday and start_week <= week_bday <= end_week:
                upcoming_count += 1
            if next_bday and start_next <= next_bday <= end_next:
                next_week_count += 1
    except Exception:
        upcoming_count = 0
        next_week_count = 0

    # outbound & inapp stats
    outbound_count = OutboundMessage.objects.count()
    sent_count = OutboundMessage.objects.filter(status__in=['sent', 'delivered']).count()
    delivered_inapp = InboxEntry.objects.count()
    read = InboxEntry.objects.filter(is_read=True).count()
    open_rate = int((read / delivered_inapp * 100)) if delivered_inapp else 0

    # Messages sent (all channels)
    messages_sent = outbound_count + InAppMessage.objects.filter(is_draft=False).count()

    # Active birthday templates
    active_birthday_templates = BirthdayTemplate.objects.filter(enabled=True, template_type=BirthdayTemplate.TEMPLATE_TYPE_BIRTHDAY).count()

    return JsonResponse({
        # existing keys (compatibility)
        'total_customers': total_customers,
        'birthdays_this_month': birthdays_month,
        'messages_sent': messages_sent,
        'open_rate_percent': open_rate,
        'clients_count': clients,
        'marketers_count': marketers,
        # new keys used by auto_birthday_templates.html
        'today_birthdays': today_birthdays_count,
        'upcoming_birthdays': upcoming_count,
        'birthday_messages_sent': messages_sent,  # using overall messages for now
        'active_birthday_templates': active_birthday_templates,
        'next_week_birthdays': next_week_count,
    }, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_birthdays_upcoming(request):
    today = lagos_today()
    period = (request.GET.get('period') or '').lower()
    # calendar week starts on Sunday (weekday() Monday=0..Sunday=6 => Sunday offset is (weekday+1)%7)
    if period == 'weekly':
        start = today - timedelta(days=(today.weekday() + 1) % 7)  # Sunday
        end = start + timedelta(days=6)
    elif period == 'next_week':
        start = (today - timedelta(days=(today.weekday() + 1) % 7)) + timedelta(days=7)
        end = start + timedelta(days=6)
    else:
        # default to rolling 7-day window from today
        start = today
        end = today + timedelta(days=7)
    # Build from concrete ClientUser and MarketerUser to mirror directories
    upcoming = []
    def add_users(qs, role_label):
        for u in qs:
            b = u.date_of_birth
            if not b:
                continue
            try:
                # Construct birthday using the window's year to avoid year-boundary errors
                this_year_bday = datetime(year=start.year, month=b.month, day=b.day).date()
            except Exception:
                continue
            if start <= this_year_bday <= end:
                # Derive names from first/last or split full_name
                first = (u.first_name or '').strip()
                last = (u.last_name or '').strip()
                if not first and not last:
                    parts = (u.full_name or '').strip().split()
                    first = parts[0] if parts else ''
                    last = ' '.join(parts[1:]) if len(parts) > 1 else ''
                upcoming.append({
                    'id': u.id,
                    'firstName': first,
                    'lastName': last,
                    'role': role_label,
                    'date': this_year_bday.isoformat(),
                    'is_today': this_year_bday == today,
                })

    add_users(ClientUser.objects.exclude(date_of_birth__isnull=True), 'client')
    add_users(MarketerUser.objects.exclude(date_of_birth__isnull=True), 'marketer')
    return JsonResponse({'birthdays': upcoming}, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_audience_options(request):
    """Return counts for audiences used by Birthday Templates dropdown.
    Clients + Marketers are counted from User.role (case-insensitive).
    Royal Elite and Estate Ambassador are derived from ClientUser.rank_tag
    computed from fully-paid transactions (via prefetch for efficiency).
    """
    try:
        clients_count = User.objects.filter(Q(role__iexact='client')).count()
        marketers_count = User.objects.filter(Q(role__iexact='marketer')).count()
        total_users = clients_count + marketers_count

        # Compute rank-based counts by evaluating rank_tag property
        royal_elite = 0
        estate_ambassador = 0
        for cu in ClientUser.with_fully_paid_prefetch():
            tag = cu.rank_tag
            if tag == 'Royal Elite':
                royal_elite += 1
            elif tag == 'Estate Ambassador':
                estate_ambassador += 1

        data = {
            'total_users': total_users,
            'clients': clients_count,
            'marketers': marketers_count,
            'royal_elite': royal_elite,
            'estate_ambassador': estate_ambassador,
            'options': [
                {'value': 'all', 'label': 'All Users', 'count': total_users},
                {'value': 'clients', 'label': 'All Clients', 'count': clients_count},
                {'value': 'marketers', 'label': 'All Marketers', 'count': marketers_count},
                {'value': 'royal_elite', 'label': 'Royal Elite Clients', 'count': royal_elite},
                {'value': 'estate_ambassador', 'label': 'Estate Ambassador Clients', 'count': estate_ambassador},
            ]
        }
        return JsonResponse(data, encoder=DjangoJSONEncoder)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def _normalize_birthday_for_year(dob, year):
    try:
        return date(year, dob.month, dob.day)
    except ValueError:
        # Handle Feb 29 on non-leap years by falling back to Feb 28
        if dob.month == 2 and dob.day == 29:
            try:
                return date(year, 2, 29)
            except ValueError:
                return date(year, 2, 28)
        raise


def _build_window_day_set(start, end):
    days = set()
    current = start
    while current <= end:
        days.add((current.month, current.day))
        current += timedelta(days=1)
    return days


def _collect_support_birthdays(request, start, end, today):
    window_days = _build_window_day_set(start, end)
    # include Feb 29 celebrants on Feb 28 in non-leap years when the window covers it
    include_feb28 = (2, 28) in window_days and not calendar.isleap(today.year)

    current_year = today.year
    marketer_rank_source = list(
        MarketerUser.objects.filter(is_deleted=False).values('id', 'full_name', 'first_name', 'last_name')
    )
    marketer_name_map = {}
    for row in marketer_rank_source:
        name = row.get('full_name') or f"{row.get('first_name') or ''} {row.get('last_name') or ''}".strip()
        marketer_name_map[row['id']] = name.strip() or 'Marketer'

    marketer_ids = list(marketer_name_map.keys())
    sales_by_marketer = {mid: 0.0 for mid in marketer_ids}

    if marketer_ids:
        perf_records = MarketerPerformanceRecord.objects.filter(
            period_type='annual',
            specific_period=str(current_year),
            marketer_id__in=marketer_ids,
        )
        for perf in perf_records:
            try:
                amount = float(perf.total_sales or 0)
            except (TypeError, ValueError):
                amount = 0.0
            sales_by_marketer[perf.marketer_id] = max(sales_by_marketer.get(perf.marketer_id, 0.0), amount)

        txn_agg = (
            Transaction.objects
            .filter(marketer_id__in=marketer_ids, transaction_date__year=current_year)
            .values('marketer')
            .annotate(
                total=Coalesce(
                    Sum('total_amount'),
                    Value(Decimal('0'), output_field=DecimalField(max_digits=18, decimal_places=2)),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )
        )
        for item in txn_agg:
            marketer_id = item.get('marketer')
            if not marketer_id:
                continue
            try:
                amount = float(item.get('total') or 0)
            except (TypeError, ValueError):
                amount = 0.0
            sales_by_marketer[marketer_id] = max(sales_by_marketer.get(marketer_id, 0.0), amount)

    if marketer_ids:
        assigned_counts = {
            row['assigned_marketer']: row['total']
            for row in (
                ClientUser.objects
                .filter(
                    assigned_marketer_id__in=marketer_ids,
                    assigned_marketer__isnull=False,
                    is_deleted=False,
                )
                .values('assigned_marketer')
                .annotate(total=Count('id'))
            )
        }
    else:
        assigned_counts = {}

    ordered_marketer_ids = sorted(
        marketer_ids,
        key=lambda mid: (
            -sales_by_marketer.get(mid, 0.0),
            (marketer_name_map.get(mid) or '').lower(),
            mid,
        ),
    )
    rank_positions = {mid: idx + 1 for idx, mid in enumerate(ordered_marketer_ids)}

    def marketer_rank_label(position: int | None) -> str:
        if not position:
            return ''
        if position <= 5:
            return 'Top 5'
        if position <= 10:
            return 'Top 10'
        if position <= 20:
            return 'Top 20'
        if position <= 50:
            return 'Top 50'
        return ''

    def iter_users(qs, role_label):
        for user in qs.exclude(date_of_birth__isnull=True):
            dob = user.date_of_birth
            if not dob:
                continue
            if (dob.month, dob.day) not in window_days:
                if not (include_feb28 and dob.month == 2 and dob.day == 29):
                    continue

            # Determine the correct calendar year for the upcoming celebration within the window
            candidate_year = start.year
            normalized = _normalize_birthday_for_year(dob, candidate_year)
            if normalized < start:
                normalized = _normalize_birthday_for_year(dob, candidate_year + 1)
            days_until = (normalized - today).days
            full_name = (user.full_name or f"{user.first_name or ''} {user.last_name or ''}").strip()
            if not full_name:
                full_name = "Unnamed"
            location = getattr(user, 'address', '') or getattr(user, 'company_profile', None)
            if location and hasattr(location, 'name'):
                location = location.name
            rank = ''
            properties_owned = None
            assigned_clients = None
            yearly_sales = None
            marketer_rank_position = None
            if role_label == 'marketer':
                assigned_clients = assigned_counts.get(user.id, 0)
                yearly_sales = round(sales_by_marketer.get(user.id, 0.0), 2)
                marketer_rank_position = rank_positions.get(user.id)
                rank_candidate = marketer_rank_label(marketer_rank_position)
                if rank_candidate:
                    rank = f"{rank_candidate} Marketer"
            else:
                try:
                    rank = getattr(user, 'rank_tag') if hasattr(user, 'rank_tag') else ''
                except Exception:
                    rank = ''
            if not rank:
                rank = 'Affiliate Partner' if role_label == 'marketer' else role_label.title()

            try:
                properties_owned = getattr(user, 'plot_count') if hasattr(user, 'plot_count') else None
            except Exception:
                properties_owned = None
            if role_label == 'marketer':
                properties_owned = None

            address_val = getattr(user, 'address', '') or ''

            avatar_url = ''
            try:
                avatar_obj = getattr(user, 'profile_image', None)
                if avatar_obj and hasattr(avatar_obj, 'url'):
                    avatar_url = avatar_obj.url or ''
            except Exception:
                avatar_url = ''

            if avatar_url and request is not None:
                try:
                    avatar_url = request.build_absolute_uri(avatar_url) if not avatar_url.lower().startswith('http') else avatar_url
                except Exception:
                    pass

            entry = {
                'id': user.id,
                'name': full_name,
                'role': role_label,
                'birthday': normalized.isoformat(),
                'originalBirthday': dob.isoformat(),
                'daysUntil': days_until,
                'email': getattr(user, 'email', None),
                'phone': getattr(user, 'phone', None),
                'location': location or '',
                'rank': rank,
                'propertiesOwned': properties_owned,
                'address': address_val,
                'avatar': avatar_url,
                'rankPosition': marketer_rank_position if role_label == 'marketer' else None,
                'assignedClients': assigned_clients if role_label == 'marketer' else None,
                'yearlySales': yearly_sales if role_label == 'marketer' else None,
            }
            yield entry

    birthdays = []
    for role, queryset in (
        ('client', ClientUser.objects.select_related('company_profile')),
        ('marketer', MarketerUser.objects.select_related('company_profile')),
    ):
        birthdays.extend(iter_users(queryset, role))

    birthdays.sort(key=lambda item: (item['daysUntil'], item['name']))
    return birthdays


@require_http_methods(['GET'])
def api_birthdays_summary(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    today = lagos_today()
    week_end = today + timedelta(days=6)
    # Remaining days in current month (from today)
    last_day = calendar.monthrange(today.year, today.month)[1]
    month_end = today.replace(day=last_day)

    today_entries = _collect_support_birthdays(request, today, today, today)
    week_entries = _collect_support_birthdays(request, today, week_end, today)
    month_entries = _collect_support_birthdays(request, today, month_end, today)

    data = {
        'generatedAt': timezone.now().isoformat(),
        'today': today_entries,
        'thisWeek': week_entries,
        'thisMonth': month_entries,
        'weekRange': {
            'start': today.isoformat(),
            'end': week_end.isoformat(),
        },
        'monthRange': {
            'start': today.isoformat(),
            'end': month_end.isoformat(),
        },
    }
    return JsonResponse(data, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
def api_birthdays_counts(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    today = lagos_today()
    week_end = today + timedelta(days=6)

    today_total = len(_collect_support_birthdays(request, today, today, today))
    week_total = len(_collect_support_birthdays(request, today, week_end, today))

    return JsonResponse({
        'today': today_total,
        'thisWeek': week_total,
    })


def _normalize_special_day_event(event):
    props = event.get('extendedProps', {}) if isinstance(event, dict) else {}
    title = event.get('title') if isinstance(event, dict) else None
    title = title or props.get('name') or props.get('title') or 'Event'
    source = (props.get('source') or '').lower()
    is_local = bool(props.get('local_db'))
    country = (props.get('countryCode') or '').upper()
    is_nigerian = country == 'NG' or source in {'ng', 'nigeria', 'nigerian'} or is_local
    event_type = (props.get('type') or ('custom' if is_local else ('national' if is_nigerian else 'custom'))).lower()
    category = (props.get('category') or event_type or 'custom').lower()
    include = is_nigerian or event_type in {'national', 'custom', 'commercial', 'observance', 'religious'} or is_local
    if not include:
        return None

    start_raw = (event.get('start') if isinstance(event, dict) else '') or ''
    date_only = start_raw.split('T')[0] if isinstance(start_raw, str) else ''

    description = props.get('description') or event.get('description') if isinstance(event, dict) else None
    if not description:
        description = title

    return {
        'id': event.get('id') if isinstance(event, dict) else None,
        'name': props.get('name') or title,
        'title': title,
        'date': date_only,
        'description': description,
        'countryCode': props.get('countryCode') or ('NG' if is_nigerian else ''),
        'type': event_type,
        'category': category,
        'source': props.get('source'),
        'isNigerian': is_nigerian,
        'isCustom': is_local or category == 'custom',
        'isRecurring': bool(props.get('recurring')),
        'localEventId': props.get('eventId'),
    }


def _collect_special_days_window(start_date, end_date):
    if end_date < start_date:
        return []

    normalized_events = []
    seen_keys = set()

    month_cursor = date(start_date.year, start_date.month, 1)
    end_month = date(end_date.year, end_date.month, 1)

    while month_cursor <= end_month:
        events = _get_special_day_events(month_cursor.year, month_cursor.month)
        for event in events:
            normalized = _normalize_special_day_event(event)
            if not normalized:
                continue
            event_date = normalized.get('date')
            if not event_date:
                continue
            try:
                event_date_obj = datetime_module.date.fromisoformat(event_date)
            except Exception:
                continue
            if not (start_date <= event_date_obj <= end_date):
                continue
            key = (event_date, (normalized.get('name') or '').lower())
            if key in seen_keys:
                continue
            seen_keys.add(key)
            normalized_events.append(normalized)

        if month_cursor.month == 12:
            month_cursor = date(month_cursor.year + 1, 1, 1)
        else:
            month_cursor = date(month_cursor.year, month_cursor.month + 1, 1)

    normalized_events.sort(key=lambda item: (item.get('date') or '', item.get('name') or ''))
    return normalized_events


@require_http_methods(['GET'])
def api_special_days_counts(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    today = lagos_today()
    month_end = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    today_events = _collect_special_days_window(today, today)
    upcoming_events = _collect_special_days_window(today + timedelta(days=1), month_end)

    today_event = today_events[0] if today_events else None
    next_event = upcoming_events[0] if upcoming_events else None

    return JsonResponse({
        'today': today_event,
        'next': next_event,
    }, encoder=DjangoJSONEncoder)


# ----------------------- Staff Management APIs -----------------------

def _serialize_staff(s):
    """Serialize a StaffMember row for the Staff Directory UI."""
    full_name = (s.full_name or '').strip()
    employment_date = s.employment_date.isoformat() if s.employment_date else None
    dob = s.date_of_birth.isoformat() if s.date_of_birth else None
    return {
        'id': s.id,
        'name': full_name,
        'full_name': full_name,
        'email': s.email or '',
        'phone': s.phone or '',
        'whatsapp': s.whatsapp or '',
        'address': s.address or '',
        'role': s.role or '',
        'employmentDate': employment_date,
        'employment_date': employment_date,
        'status': 'active' if s.active else 'inactive',
        'active': bool(s.active),
        'date_of_birth': dob,
        'dateOfBirth': dob,
    }


def _get_staff_payload(request):
    """Return request payload as dict for staff create/update operations."""
    if request.content_type and 'application/json' in request.content_type:
        try:
            raw = request.body.decode('utf-8') if request.body else '{}'
            data = json.loads(raw)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return request.POST or {}


def _staff_form_values(payload):
    """Normalize incoming payload values for staff forms."""

    def _get_value(keys, *, to_lower=False):
        for key in keys:
            if key in payload:
                val = payload.get(key)
                if isinstance(val, str):
                    val = val.strip()
                    if to_lower:
                        val = val.lower()
                return val
        return ''

    employment_raw = payload.get('employment_date') or payload.get('employmentDate')
    dob_raw = payload.get('date_of_birth') or payload.get('dateOfBirth')

    active_raw = payload.get('active')
    if isinstance(active_raw, str):
        active_normalized = active_raw.strip().lower()
        if active_normalized in ('true', '1', 'yes', 'on', 'active'):
            active_val = True
        elif active_normalized in ('false', '0', 'no', 'off', 'inactive'):
            active_val = False
        else:
            active_val = None
    elif isinstance(active_raw, bool):
        active_val = active_raw
    else:
        active_val = None

    return {
        'full_name': _get_value(['full_name', 'name']),
        'email': _get_value(['email'], to_lower=True),
        'phone': _get_value(['phone_number', 'phone']),
        'whatsapp': _get_value(['whatsapp']),
        'address': _get_value(['address']),
        'role': _get_value(['role', 'job']),
        'employment_date': _parse_iso_date(employment_raw),
        'date_of_birth': _parse_iso_date(dob_raw),
        'active': active_val,
    }


@require_http_methods(['POST'])
@login_required
@support_required
def api_staff_create(request):
    payload = _get_staff_payload(request)
    values = _staff_form_values(payload)

    required_map = {
        'full_name': 'Full name',
        'email': 'Email address',
        'role': 'Role/Position',
        'employment_date': 'Employment date',
        'address': 'Address',
        'phone': 'Phone number',
        'date_of_birth': 'Date of birth',
    }

    missing = [label for key, label in required_map.items() if not values.get(key)]
    if missing:
        return JsonResponse({'ok': False, 'error': f"Missing required fields: {', '.join(missing)}"}, status=400)

    staff = StaffMember.objects.filter(email=values['email']).first()
    created = False
    reactivated = False

    if staff:
        updated_fields = []
        for field in ['full_name', 'phone', 'whatsapp', 'address', 'role', 'employment_date', 'date_of_birth']:
            new_value = values.get(field)
            if new_value and getattr(staff, field) != new_value:
                setattr(staff, field, new_value)
                updated_fields.append(field)
        if not staff.active:
            staff.active = True
            reactivated = True
            updated_fields.append('active')

        if not updated_fields:
            return JsonResponse({'ok': False, 'error': 'Staff member with this email already exists.'}, status=400)

        staff.save(update_fields=updated_fields)
        action = "Reactivated" if reactivated else "Updated"
        ActivityLog.objects.create(actor=request.user, action=f"{action} staff {staff.full_name} ({staff.email})")
    else:
        staff = StaffMember.objects.create(
            full_name=values['full_name'],
            email=values['email'],
            phone=values['phone'],
            whatsapp=values.get('whatsapp') or '',
            address=values['address'],
            role=values['role'],
            employment_date=values['employment_date'],
            date_of_birth=values['date_of_birth'],
            active=True if values.get('active') is None else bool(values['active']),
            created_by=request.user,
        )
        created = True
        ActivityLog.objects.create(actor=request.user, action=f"Added staff {staff.full_name} ({staff.email})")

    return JsonResponse({
        'ok': True,
        'created': created,
        'reactivated': reactivated,
        'staff': _serialize_staff(staff),
    }, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_staff_stats(request):
    total = StaffMember.objects.count()
    active = StaffMember.objects.filter(active=True).count()

    staff_items = OutboundMessageItem.objects.filter(outbound__message_type='staff')
    delivered = staff_items.filter(status__in=['sent', 'delivered']).count()
    failed = staff_items.filter(status='failed').count()
    return JsonResponse({
        'total': total,
        'active': active,
        'messages_delivered': delivered,
        'messages_failed': failed,
    })


@require_http_methods(['GET'])
@login_required
@support_required
def api_staff_list(request):
    q = (request.GET.get('q') or '').strip().lower()
    filter_role = (request.GET.get('role') or '').strip().lower()
    data = []
    qs = StaffMember.objects.filter(active=True)
    if q:
        qs = qs.filter(
            Q(full_name__icontains=q) | Q(email__icontains=q) | Q(phone__icontains=q) | Q(role__icontains=q)
        )
    if filter_role:
        qs = qs.filter(role__icontains=filter_role)
    for s in qs.order_by('-created_at')[:500]:
        data.append(_serialize_staff(s))
    return JsonResponse({'results': data})


@require_http_methods(['GET'])
@login_required
@support_required
def api_staff_detail(request, staff_id):
    staff = get_object_or_404(StaffMember, id=staff_id)
    return JsonResponse({'ok': True, 'staff': _serialize_staff(staff)})


@require_http_methods(['GET'])
@login_required
@support_required
def api_staff_former(request):
    data = []
    for s in StaffMember.objects.filter(active=False).order_by('-created_at')[:500]:
        data.append(_serialize_staff(s))
    return JsonResponse({'results': data})


def _serialize_failed_staff_item(item: OutboundMessageItem):
    recipient = item.recipient
    outbound = item.outbound
    return {
        'id': str(item.id),
        'messageId': str(outbound.id) if outbound else None,
        'recipientId': recipient.id if recipient else None,
        'recipientName': recipient.get_full_name() if recipient else None,
        'recipientEmail': getattr(recipient, 'email', None),
        'recipientPhone': getattr(recipient, 'phone', None),
        'channel': item.channel,
        'error': item.error,
        'failedAt': item.created_at.isoformat() if item.created_at else None,
        'retryCount': getattr(outbound, 'retry_count', 0) if outbound else 0,
    }


def _retry_outbound_item(item: OutboundMessageItem):
    recipient = item.recipient
    if not recipient:
        item.error = 'No recipient'
        item.save(update_fields=['error'])
        return False

    subject = item.subject or ''
    body = item.body or ''

    if item.channel == 'email':
        ok, err = send_email_message(subject or 'Notification', body, getattr(recipient, 'email', None))
    elif item.channel == 'sms':
        ok, err = send_sms_message(body, getattr(recipient, 'phone', None))
    else:
        ok, err = send_inapp_message(body or subject, recipient)

    if ok:
        item.status = 'sent'
        item.sent_at = timezone.now()
        item.error = ''
        item.save(update_fields=['status', 'sent_at', 'error'])
        return True

    item.error = str(err) if err else 'Unknown error'
    item.save(update_fields=['error'])
    return False


@require_http_methods(["GET"])
@login_required
@support_required
def api_staff_failed_messages(request):
    qs = OutboundMessageItem.objects.select_related('recipient', 'outbound') \
        .filter(status='failed', outbound__message_type='staff').order_by('-created_at')[:200]
    data = [_serialize_failed_staff_item(item) for item in qs]
    return JsonResponse({'results': data}, encoder=DjangoJSONEncoder)


@require_http_methods(["POST"])
@login_required
@support_required
def api_staff_failed_retry(request):
    message_id = request.POST.get('message_id')
    if not message_id and request.body:
        try:
            message_id = json.loads(request.body.decode('utf-8')).get('message_id')
        except Exception:
            message_id = None
    if not message_id:
        return JsonResponse({'success': False, 'error': 'message_id is required'}, status=400)

    item = get_object_or_404(OutboundMessageItem.objects.select_related('outbound', 'recipient'), id=message_id)
    if item.outbound.message_type != 'staff':
        return JsonResponse({'success': False, 'error': 'Message is not a staff automation'}, status=400)

    success = _retry_outbound_item(item)
    if success:
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': item.error or 'Retry failed'}, status=400)


@require_http_methods(["POST"])
@login_required
@support_required
def api_staff_failed_retry_all(request):
    qs = OutboundMessageItem.objects.select_related('outbound', 'recipient') \
        .filter(status='failed', outbound__message_type='staff')

    count = 0
    success_count = 0
    for item in qs.iterator():
        count += 1
        if _retry_outbound_item(item):
            success_count += 1

    return JsonResponse({'success': True, 'count': count, 'retried': success_count})


@require_http_methods(["POST"])
@login_required
@support_required
def api_staff_failed_delete(request):
    message_id = request.POST.get('message_id')
    if not message_id and request.body:
        try:
            message_id = json.loads(request.body.decode('utf-8')).get('message_id')
        except Exception:
            message_id = None
    if not message_id:
        return JsonResponse({'success': False, 'error': 'message_id is required'}, status=400)

    item = get_object_or_404(OutboundMessageItem, id=message_id, outbound__message_type='staff')
    item.delete()
    return JsonResponse({'success': True})


def _parse_iso_date(val):
    try:
        if not val:
            return None
        return datetime.fromisoformat(val).date()
    except Exception:
        try:
            return datetime.strptime(val, '%Y-%m-%d').date()
        except Exception:
            return None


@require_http_methods(['POST'])
@login_required
@support_required
@require_http_methods(['POST'])
@login_required
@support_required
def api_staff_remove(request):
    payload = None
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        payload = request.POST or {}
    staff_id = payload.get('user_id') or payload.get('staffId') or payload.get('id')
    if not staff_id:
        return JsonResponse({'error': 'id required'}, status=400)
    s = get_object_or_404(StaffMember, id=staff_id)
    s.active = False
    s.save(update_fields=['active'])
    reasons = payload.get('reasons') or []
    notes = payload.get('notes') or ''
    ActivityLog.objects.create(actor=request.user, action=f"Removed staff {s.full_name} reasons={reasons} notes={(notes or '')[:120]}")
    return JsonResponse({'ok': True})


@require_http_methods(['POST'])
@login_required
@support_required
def api_staff_reactivate(request):
    payload = None
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        payload = request.POST or {}
    staff_id = payload.get('user_id') or payload.get('staffId') or payload.get('id')
    if not staff_id:
        return JsonResponse({'error': 'id required'}, status=400)
    s = get_object_or_404(StaffMember, id=staff_id)
    s.active = True
    s.save(update_fields=['active'])
    ActivityLog.objects.create(actor=request.user, action=f"Reactivated staff {s.full_name}")
    return JsonResponse({'ok': True})


@require_http_methods(['POST'])
@login_required
@support_required
def api_staff_import(request):
    f = request.FILES.get('file')
    if not f:
        return JsonResponse({'error': 'No file provided'}, status=400)
    name = f.name.lower()
    created = 0
    updated = 0
    errors = []

    def upsert(row):
        nonlocal created, updated
        try:
            email = (row.get('email') or row.get('Email') or '').strip().lower()
            full_name = (row.get('full_name') or row.get('Full Name') or row.get('name') or '').strip()
            phone = (row.get('phone') or row.get('Phone') or '').strip()
            role = (row.get('role') or row.get('Role') or row.get('job') or '').strip()
            address = (row.get('address') or row.get('Address') or '').strip()
            dob = _parse_iso_date(row.get('date_of_birth') or row.get('Date of Birth'))
            employment_date = _parse_iso_date(row.get('employment_date') or row.get('Employment Date'))
            if not email or not full_name:
                return
            obj, was_created = StaffMember.objects.get_or_create(
                email=email,
                defaults={
                    'full_name': full_name,
                    'phone': phone,
                    'address': address,
                    'role': role,
                    'employment_date': employment_date,
                    'date_of_birth': dob,
                    'active': True,
                    'created_by': request.user,
                }
            )
            if was_created:
                created += 1
            else:
                # update existing
                obj.full_name = full_name or obj.full_name
                obj.phone = phone or obj.phone
                obj.address = address or obj.address
                obj.role = role or obj.role
                obj.employment_date = employment_date or obj.employment_date
                obj.date_of_birth = dob or obj.date_of_birth
                obj.active = True
                obj.save()
                updated += 1
        except Exception as e:
            errors.append(str(e))

    if name.endswith('.csv'):
        try:
            # Ensure text mode wrapper with utf-8
            text = TextIOWrapper(f.file, encoding='utf-8')
            reader = csv.DictReader(text)
            for row in reader:
                upsert(row)
        except Exception as e:
            return JsonResponse({'error': f'CSV parse failed: {e}'}, status=400)
    elif name.endswith('.xlsx') or name.endswith('.xlsm'):
        if not OPENPYXL_AVAILABLE:
            return JsonResponse({'error': 'Excel support requires openpyxl to be installed'}, status=400)
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for r in ws.iter_rows(min_row=2):
                row = {headers[i]: (r[i].value if i < len(r) else None) for i in range(len(headers))}
                upsert(row)
        except Exception as e:
            return JsonResponse({'error': f'Excel parse failed: {e}'}, status=400)
    else:
        return JsonResponse({'error': 'Unsupported file type. Upload .csv or .xlsx'}, status=400)

    ActivityLog.objects.create(actor=request.user, action=f"Imported staff: +{created}/~{updated} updates")
    return JsonResponse({'ok': True, 'created': created, 'updated': updated, 'errors': errors})


@require_http_methods(['GET'])
@login_required
@support_required
def api_messages_status(request):
    items = []
    # Show only staff-targeted In-App messages (sent by current user)
    for msg in InAppMessage.objects.filter(is_draft=False, sender=request.user).order_by('-created_at')[:100]:
        items.append({
            'id': f"inapp-{msg.id}",
            'subject': msg.subject or '(No subject)',
            'recipients': msg.recipients.count(),
            'channel': 'In-App',
            'status': 'delivered',
            'sentAt': (msg.created_at.strftime('%Y-%m-%d %H:%M') if msg.created_at else ''),
            'successRate': 100,
        })
    # sort by sentAt desc if present
    try:
        items.sort(key=lambda x: x.get('sentAt') or '', reverse=True)
    except Exception:
        pass
    return JsonResponse({'results': items})


@require_http_methods(['POST'])
@login_required
@support_required
def api_run_scheduled_triggers(request):
    now = timezone.now()
    triggers = ScheduledTrigger.objects.filter(enabled=True)
    created = []
    for t in triggers:
        tpl = t.template
        should_run = False

        if not t.daily and t.one_shot_at:
            # run one-shot if scheduled time passed and not yet run
            if t.one_shot_at <= now and (not t.last_run or t.last_run < t.one_shot_at):
                should_run = True

        if t.daily and t.time_of_day:
            # compute today's scheduled datetime in current tz
            today_scheduled = now.replace(hour=t.time_of_day.hour, minute=t.time_of_day.minute, second=0, microsecond=0)
            if today_scheduled <= now and (not t.last_run or t.last_run.date() < now.date()):
                should_run = True

        if should_run:
            if tpl.audience == 'clients':
                recipients_qs = User.objects.filter(role='client')
            elif tpl.audience == 'marketers':
                recipients_qs = User.objects.filter(role='marketer')
            else:
                recipients_qs = User.objects.all()

            recipients_count = recipients_qs.count()
            sample = list(recipients_qs.values_list('first_name', 'last_name')[:3])
            sample_strs = [f"{a or ''} {b or ''}".strip() for a, b in sample]

            om = OutboundMessage.objects.create(
                channel=tpl.channel,
                audience=tpl.audience,
                subject=tpl.subject,
                body=tpl.body,
                created_by=None,
                recipients_count=recipients_count,
                recipients_sample=sample_strs,
                scheduled_for=now,
                status='queued'
            )
            t.last_run = now
            t.save(update_fields=['last_run'])
            ActivityLog.objects.create(actor=None, action=f"Scheduled trigger queued outbound id={om.id} for template {tpl.name}")
            created.append(str(om.id))

    return JsonResponse({'ok': True, 'created': created}, encoder=DjangoJSONEncoder)


# Conversation endpoints
@require_http_methods(['POST'])
@login_required
@support_required
def api_conversation_create(request):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    part_ids = payload.get('participant_ids', [])
    subject = payload.get('subject', None)

    conv = SupportConversation.objects.create(subject=subject, created_by=request.user)
    if part_ids:
        participants = User.objects.filter(id__in=part_ids)
        if participants.exists():
            conv.participants.set(participants)
    conv.participants.add(request.user)
    ActivityLog.objects.create(actor=request.user, action=f"Created conversation {conv.id} ({subject})")
    return JsonResponse({'ok': True, 'id': conv.id}, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_conversation_list(request):
    qs = SupportConversation.objects.filter(participants=request.user).order_by('-created_at')[:200]
    data = []
    for c in qs:
        data.append({
            'id': c.id,
            'subject': c.subject,
            'participants': [{'id': u.id, 'name': u.get_full_name(), 'role': getattr(u, 'role', '')} for u in c.participants.all()],
            'createdAt': c.created_at.isoformat() if c.created_at else None
        })
    return JsonResponse({'conversations': data}, encoder=DjangoJSONEncoder)


@require_http_methods(['GET'])
@login_required
@support_required
def api_conversation_messages(request, conv_id):
    conv = get_object_or_404(SupportConversation, id=conv_id)
    if not conv.participants.filter(id=request.user.id).exists():
        return HttpResponseForbidden("Not a participant")

    msgs = conv.messages.select_related('sender').all().order_by('created_at')
    data = []
    for m in msgs:
        data.append({
            'id': m.id,
            'sender': {'id': m.sender.id if m.sender else None, 'name': m.sender.get_full_name() if m.sender else 'System'},
            'body': m.body,
            'createdAt': m.created_at.isoformat() if m.created_at else None,
            'isSystem': m.is_system,
            'attachments': m.attachments or []
        })
    return JsonResponse({'messages': data}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_conversation_send_message(request, conv_id):
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    conv = get_object_or_404(SupportConversation, id=conv_id)
    if not conv.participants.filter(id=request.user.id).exists():
        return HttpResponseForbidden("Not a participant")

    body = payload.get('body', '').strip()
    attachments = payload.get('attachments', [])

    if not body and not attachments:
        return HttpResponseBadRequest("Empty message")

    msg = SupportMessage.objects.create(conversation=conv, sender=request.user, body=body, attachments=attachments)
    receipts = []
    for u in conv.participants.exclude(id=request.user.id):
        receipts.append(SupportMessageReceipt(message=msg, user=u))
    SupportMessageReceipt.objects.bulk_create(receipts, ignore_conflicts=True)
    ActivityLog.objects.create(actor=request.user, action=f"Sent message in conv {conv.id}")
    return JsonResponse({'ok': True, 'id': msg.id}, encoder=DjangoJSONEncoder)



# BIRTHDAY AUTO TEMPLATE
def parse_json_body(request):
    try:
        if not request.body:
            return {}
        return json.loads(request.body.decode('utf-8'))
    except Exception:
        return None

# List and create templates
@login_required
@support_required
@require_http_methods(["GET", "POST"])
def api_templates_list_create(request):
    if request.method == "GET":
        qs = BirthdayTemplate.objects.all()
        templates = []
        for t in qs:
            templates.append({
                "id": str(t.id),
                "name": t.name,
                "channel": t.channel,
                "audience": t.audience,
                "subject": t.subject,
                "body": t.message,
                "time": t.time.isoformat() if t.time else None,
                "enabled": t.enabled,
                "created_by": getattr(t.created_by, 'username', None),
                "created_at": t.created_at.isoformat() if t.created_at else None,
                "updated_at": t.updated_at.isoformat() if t.updated_at else None,
            })
        return JsonResponse({"results": templates})

    # POST -> create
    # Accept JSON or form-data
    body = parse_json_body(request)
    if body is None:
        # fallback to form data
        body = request.POST or {}
    channel = body.get("channel", request.POST.get("channel", "sms"))
    name = body.get("name", request.POST.get("name", "")) or ""
    message = body.get("message") or body.get("body") or request.POST.get("message") or request.POST.get("body") or ""
    subject = body.get("subject") or request.POST.get("subject")
    audience = body.get("audience", request.POST.get("audience", "clients"))
    time_val = body.get("time", request.POST.get("time"))
    enabled = bool(body.get("enabled", request.POST.get("enabled", True)))

    if not message:
        return JsonResponse({"error": "message is required"}, status=400)

    tpl = BirthdayTemplate.objects.create(
        name=name,
        channel=channel,
        audience=audience,
        subject=subject,
        message=message,
        enabled=enabled,
        created_by=request.user
    )
    
    if time_val:
        try:
            from datetime import time as dt_time
            parts = time_val.split(':')
            if len(parts) >= 2:
                hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                tpl.time = dt_time(hour, minute, second)
                tpl.save()
        except Exception:
            pass

    return JsonResponse({
        "id": str(tpl.id),
        "name": tpl.name,
        "channel": tpl.channel,
        "audience": tpl.audience,
        "subject": tpl.subject,
        "body": tpl.message,
        "time": tpl.time.isoformat() if tpl.time else None,
        "enabled": tpl.enabled,
        "created_at": tpl.created_at.isoformat() if tpl.created_at else None,
        "updated_at": tpl.updated_at.isoformat() if tpl.updated_at else None,
    }, status=201)


# detail: GET, PATCH, DELETE
@login_required
@support_required
@require_http_methods(["GET", "PATCH", "PUT", "DELETE"])
def api_template_detail(request, tpl_id):
    tpl = get_object_or_404(BirthdayTemplate, id=tpl_id)

    if request.method == "GET":
        return JsonResponse({
            "id": str(tpl.id),
            "name": tpl.name,
            "channel": tpl.channel,
            "audience": tpl.audience,
            "subject": tpl.subject,
            "body": tpl.message,
            "time": tpl.time.isoformat() if tpl.time else None,
            "enabled": tpl.enabled,
        })

    if request.method in ("PATCH", "PUT"):
        body = parse_json_body(request)
        if body is None:
            # allow form-data too
            body = request.POST or {}
        # allow partial updates
        if "name" in body:
            tpl.name = body.get("name") or tpl.name
        if "channel" in body:
            tpl.channel = body.get("channel") or tpl.channel
        if "audience" in body:
            tpl.audience = body.get("audience") or tpl.audience
        if "message" in body or "body" in body:
            tpl.message = body.get("message") or body.get("body") or tpl.message
        if "subject" in body:
            tpl.subject = body.get("subject") or tpl.subject
        if "enabled" in body:
            tpl.enabled = bool(body.get("enabled"))
        if "time" in body:
            t = body.get("time")
            if t:
                try:
                    from datetime import time as dt_time
                    parts = t.split(':')
                    hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                    tpl.time = dt_time(hour, minute, second)
                except Exception:
                    pass
            else:
                tpl.time = None
        tpl.save()
        return JsonResponse({
            "id": str(tpl.id),
            "name": tpl.name,
            "channel": tpl.channel,
            "audience": tpl.audience,
            "subject": tpl.subject,
            "body": tpl.message,
            "time": tpl.time.isoformat() if tpl.time else None,
            "enabled": tpl.enabled,
        })

    # DELETE
    tpl.delete()
    return HttpResponse(status=204)


# auto-birthday setting GET/POST
@staff_member_required
@require_http_methods(["GET", "POST"])
def api_auto_birthday(request):
    settings_obj, _ = MessagingSettings.objects.get_or_create(pk=1)

    if request.method == "GET":
        payload = {
            "enabled": bool(settings_obj.auto_birthday),
            "send_time": settings_obj.birthday_time.isoformat() if settings_obj.birthday_time else None,
            "default_channel": settings_obj.birthday_channel,
            "days_advance": settings_obj.birthday_days_advance,
        }
        return JsonResponse(payload)

    body = parse_json_body(request)
    if body is None:
        return HttpResponseBadRequest("Invalid JSON")

    if "enabled" in body:
        settings_obj.auto_birthday = bool(body.get("enabled"))

    if "default_channel" in body:
        channel = body.get("default_channel") or settings_obj.birthday_channel
        if channel not in dict(MessagingSettings._meta.get_field("birthday_channel").choices):
            return JsonResponse({"error": "Invalid channel"}, status=400)
        settings_obj.birthday_channel = channel

    if "send_time" in body:
        send_time_val = body.get("send_time")
        if send_time_val:
            try:
                if isinstance(send_time_val, str):
                    hour, minute, *rest = [int(x) for x in send_time_val.split(":")]
                    second = rest[0] if rest else 0
                    from datetime import time as dt_time
                    settings_obj.birthday_time = dt_time(hour, minute, second)
                else:
                    return JsonResponse({"error": "send_time must be HH:MM or HH:MM:SS"}, status=400)
            except Exception:
                return JsonResponse({"error": "Invalid send_time"}, status=400)
        else:
            settings_obj.birthday_time = None

    if "days_advance" in body:
        try:
            days_val = int(body.get("days_advance"))
            if days_val < 0:
                return JsonResponse({"error": "days_advance must be >= 0"}, status=400)
            settings_obj.birthday_days_advance = days_val
        except (TypeError, ValueError):
            return JsonResponse({"error": "days_advance must be an integer"}, status=400)

    settings_obj.save(update_fields=[
        "auto_birthday",
        "birthday_channel",
        "birthday_time",
        "birthday_days_advance",
        "updated_at",
    ])

    response = {
        "enabled": bool(settings_obj.auto_birthday),
        "send_time": settings_obj.birthday_time.isoformat() if settings_obj.birthday_time else None,
        "default_channel": settings_obj.birthday_channel,
        "days_advance": settings_obj.birthday_days_advance,
    }
    return JsonResponse(response)


def _serialize_failed_birthday_item(item: OutboundMessageItem):
    recipient = item.recipient
    outbound = item.outbound
    birthday = getattr(recipient, 'date_of_birth', None)
    return {
        'id': str(item.id),
        'messageId': str(outbound.id) if outbound else None,
        'recipientId': recipient.id if recipient else None,
        'recipientName': ((recipient.first_name or '') + ' ' + (recipient.last_name or '')).strip() if recipient else None,
        'recipientEmail': getattr(recipient, 'email', None),
        'recipientPhone': getattr(recipient, 'phone', None),
        'recipientRole': getattr(recipient, 'role', None),
        'recipientBirthday': birthday.isoformat() if birthday else None,
        'channel': item.channel,
        'error': item.error,
        'failedAt': item.created_at.isoformat() if item.created_at else None,
        'retryCount': getattr(outbound, 'retry_count', 0) if outbound else 0,
    }


@require_http_methods(["GET"])
@login_required
@support_required
def api_birthday_failed(request):
    qs = OutboundMessageItem.objects.select_related('recipient', 'outbound')\
        .filter(status='failed', outbound__message_type='birthday')

    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    if start_date:
        try:
            qs = qs.filter(created_at__date__gte=datetime_module.date.fromisoformat(start_date))
        except Exception:
            pass
    if end_date:
        try:
            qs = qs.filter(created_at__date__lte=datetime_module.date.fromisoformat(end_date))
        except Exception:
            pass

    qs = qs.order_by('-created_at')[:200]
    data = [_serialize_failed_birthday_item(item) for item in qs]
    return JsonResponse({'results': data}, encoder=DjangoJSONEncoder)


def _retry_birthday_item(item: OutboundMessageItem):
    recipient = item.recipient
    if not recipient:
        item.error = 'No recipient'
        item.save(update_fields=['error'])
        return False

    subject = item.subject or 'Happy Birthday'
    body = item.body or ''

    if item.channel == 'email':
        ok, err = send_email_message(subject, body, getattr(recipient, 'email', None))
    elif item.channel == 'sms':
        ok, err = send_sms_message(body, getattr(recipient, 'phone', None))
    else:
        ok, err = send_inapp_message(body, recipient)

    if ok:
        item.status = 'sent'
        item.sent_at = timezone.now()
        item.error = ''
        item.save(update_fields=['status', 'sent_at', 'error'])
        return True

    item.error = str(err) if err else 'Unknown error'
    item.save(update_fields=['error'])
    return False


@require_http_methods(["POST"])
@login_required
@support_required
def api_birthday_retry(request):
    message_id = request.POST.get('message_id') or (json.loads(request.body.decode('utf-8')).get('message_id') if request.body else None)
    if not message_id:
        return JsonResponse({'success': False, 'error': 'message_id is required'}, status=400)

    item = get_object_or_404(OutboundMessageItem.objects.select_related('outbound', 'recipient'), id=message_id)

    success = _retry_birthday_item(item)
    if success:
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'error': item.error or 'Retry failed'}, status=400)


@require_http_methods(["POST"])
@login_required
@support_required
def api_birthday_retry_all(request):
    qs = OutboundMessageItem.objects.select_related('outbound', 'recipient')\
        .filter(status='failed', outbound__message_type='birthday')

    count = 0
    success = 0
    for item in qs.iterator():
        count += 1
        if _retry_birthday_item(item):
            success += 1

    return JsonResponse({'success': True, 'count': count, 'retried': success})


@require_http_methods(["POST"])
@login_required
@support_required
def api_birthday_delete_failed(request):
    message_id = request.POST.get('message_id') or (json.loads(request.body.decode('utf-8')).get('message_id') if request.body else None)
    if not message_id:
        return JsonResponse({'success': False, 'error': 'message_id is required'}, status=400)

    item = get_object_or_404(OutboundMessageItem, id=message_id)
    item.delete()
    return JsonResponse({'success': True})


# ----------------- Run scheduled triggers (manual endpoint) -----------------
@staff_member_required
@require_http_methods(["POST"])
def api_run_triggers(request):
    """
    Trigger the scheduled run (useful for manual tests).
    This will call the same worker code used by the management command.
    """
    # call helper
    from .tasks import run_send_birthday_messages
    result = run_send_birthday_messages()
    # result is a dict with counts
    return JsonResponse({'ok': True, 'result': result})



@require_http_methods(['GET', 'POST'])
@login_required
@support_required
def api_special_templates_list_create(request):
    """
    GET -> lists special-day templates only
    POST -> create a new special-day template
    """
    if request.method == 'GET':
        qs = BirthdayTemplate.objects.filter(template_type=BirthdayTemplate.TEMPLATE_TYPE_SPECIAL).order_by('-created_at')[:500]
        templates = []
        for t in qs:
            templates.append({
                "id": str(t.id),
                "name": t.name,
                "channel": t.channel,
                "audience": t.audience,
                "subject": t.subject,
                "message": t.message,
                "time": t.time.isoformat() if t.time else None,
                "enabled": t.enabled,
                "created_at": t.created_at.isoformat() if t.created_at else None,
                "updated_at": t.updated_at.isoformat() if t.updated_at else None,
            })
        return JsonResponse({"templates": templates}, encoder=DjangoJSONEncoder)

    # POST -> create
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    name = payload.get('name', '') or ''
    channel = payload.get('channel', 'sms')
    audience = payload.get('audience', 'clients')
    message = payload.get('message', '')
    subject = payload.get('subject', None)
    time_val = payload.get('time', None)
    enabled = bool(payload.get('enabled', True))

    if not message:
        return JsonResponse({"error": "message is required"}, status=400)

    tpl = BirthdayTemplate.objects.create(
        name=name,
        channel=channel,
        audience=audience,
        subject=subject,
        message=message,
        enabled=enabled,
        created_by=request.user,
        template_type=BirthdayTemplate.TEMPLATE_TYPE_SPECIAL
    )

    if time_val:
        try:
            from datetime import time as dt_time
            parts = str(time_val).split(':')
            if len(parts) >= 2:
                hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                tpl.time = dt_time(hour, minute, second)
                tpl.save(update_fields=['time'])
        except Exception:
            pass

    return JsonResponse({
        "id": str(tpl.id),
        "name": tpl.name,
        "channel": tpl.channel,
        "audience": tpl.audience,
        "message": tpl.message,
        "time": tpl.time.isoformat() if tpl.time else None,
        "enabled": tpl.enabled,
    }, status=201, encoder=DjangoJSONEncoder)


@require_http_methods(['GET', 'PATCH', 'PUT', 'DELETE'])
@login_required
@support_required
def api_special_template_detail(request, tpl_id):
    tpl = get_object_or_404(BirthdayTemplate, id=tpl_id, template_type=BirthdayTemplate.TEMPLATE_TYPE_SPECIAL)
    if request.method == 'GET':
        return JsonResponse({
            "id": str(tpl.id),
            "name": tpl.name,
            "channel": tpl.channel,
            "audience": tpl.audience,
            "subject": tpl.subject,
            "message": tpl.message,
            "time": tpl.time.isoformat() if tpl.time else None,
            "enabled": tpl.enabled,
        }, encoder=DjangoJSONEncoder)

    if request.method in ('PATCH', 'PUT'):
        try:
            payload = json.loads(request.body.decode('utf-8')) if request.body else {}
        except Exception:
            return HttpResponseBadRequest("Invalid JSON")

        if "name" in payload:
            tpl.name = payload.get("name") or tpl.name
        if "channel" in payload:
            tpl.channel = payload.get("channel") or tpl.channel
        if "audience" in payload:
            tpl.audience = payload.get("audience") or tpl.audience
        if "subject" in payload:
            tpl.subject = payload.get("subject") or tpl.subject
        if "message" in payload:
            tpl.message = payload.get("message") or tpl.message
        if "enabled" in payload:
            tpl.enabled = bool(payload.get("enabled"))
        if "time" in payload:
            t = payload.get("time")
            if t:
                try:
                    from datetime import time as dt_time
                    parts = t.split(':')
                    hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                    tpl.time = dt_time(hour, minute, second)
                except Exception:
                    pass
            else:
                tpl.time = None
        tpl.save()
        return JsonResponse({
            "id": str(tpl.id),
            "name": tpl.name,
            "channel": tpl.channel,
            "audience": tpl.audience,
            "subject": tpl.subject,
            "message": tpl.message,
            "time": tpl.time.isoformat() if tpl.time else None,
            "enabled": tpl.enabled,
        }, encoder=DjangoJSONEncoder)

    # DELETE
    tpl.delete()
    return HttpResponse(status=204)


@staff_member_required
@require_http_methods(['GET', 'POST'])
def api_auto_special_day(request):
    settings_obj, _ = AutoSpecialSetting.objects.get_or_create(pk=1)
    if request.method == 'GET':
        return JsonResponse({"auto_special_day": bool(settings_obj.enabled)})
    # POST -> update
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")
    if "auto_special_day" not in payload:
        return JsonResponse({"error": "auto_special_day is required"}, status=400)
    settings_obj.enabled = bool(payload.get("auto_special_day"))
    settings_obj.save(update_fields=['enabled', 'updated_at'])
    return JsonResponse({"auto_special_day": bool(settings_obj.enabled)})


def _serialize_custom_special_day(day: CustomSpecialDay) -> dict:
    return {
        'id': str(day.id),
        'name': day.name,
        'date': day.date.isoformat(),
        'description': day.description or '',
        'category': day.category or 'custom',
        'countryCode': (day.country_code or '').upper() or 'NG',
        'isRecurring': bool(day.is_recurring),
        'createdAt': day.created_at.isoformat() if day.created_at else None,
        'updatedAt': day.updated_at.isoformat() if day.updated_at else None,
    }


@require_http_methods(['GET', 'POST'])
@csrf_exempt
def api_custom_special_days(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    if request.method == 'GET':
        day_qs = CustomSpecialDay.objects.order_by('date', 'name')
        payload = [_serialize_custom_special_day(day) for day in day_qs]
        return JsonResponse({'customDays': payload}, encoder=DjangoJSONEncoder)

    body = parse_json_body(request)
    if body is None:
        return HttpResponseBadRequest('Invalid JSON payload')

    name = (body.get('name') or '').strip()
    date_str = (body.get('date') or '').strip()
    description = (body.get('description') or '').strip()
    category = (body.get('category') or 'custom').strip() or 'custom'
    country_code = (body.get('countryCode') or body.get('country_code') or 'NG').strip().upper() or 'NG'
    is_recurring = bool(body.get('isRecurring') or body.get('is_recurring'))

    if not name:
        return JsonResponse({'error': 'name is required'}, status=400)
    if not date_str:
        return JsonResponse({'error': 'date is required'}, status=400)

    try:
        parsed_date = datetime_module.datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return JsonResponse({'error': 'date must be in YYYY-MM-DD format'}, status=400)

    custom_day = CustomSpecialDay.objects.create(
        name=name,
        date=parsed_date,
        description=description,
        category=category,
        country_code=country_code,
        is_recurring=is_recurring,
        created_by=support_user if getattr(support_user, 'is_authenticated', False) else None,
    )

    return JsonResponse({'customDay': _serialize_custom_special_day(custom_day)}, status=201, encoder=DjangoJSONEncoder)


@require_http_methods(['PATCH', 'PUT', 'DELETE'])
@csrf_exempt
def api_custom_special_day_detail(request, day_id):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    custom_day = get_object_or_404(CustomSpecialDay, id=day_id)

    if request.method == 'DELETE':
        custom_day.delete()
        return HttpResponse(status=204)

    body = parse_json_body(request)
    if body is None:
        return HttpResponseBadRequest('Invalid JSON payload')

    name = body.get('name')
    date_str = body.get('date')
    description = body.get('description')
    category = body.get('category')
    country_code = body.get('countryCode') or body.get('country_code')
    is_recurring = body.get('isRecurring') if 'isRecurring' in body else body.get('is_recurring')

    if name is not None:
        name_clean = name.strip()
        if not name_clean:
            return JsonResponse({'error': 'name cannot be empty'}, status=400)
        custom_day.name = name_clean

    if date_str is not None:
        try:
            parsed_date = datetime_module.datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
            custom_day.date = parsed_date
        except Exception:
            return JsonResponse({'error': 'date must be in YYYY-MM-DD format'}, status=400)

    if description is not None:
        custom_day.description = description.strip()

    if category is not None:
        custom_day.category = category.strip() or custom_day.category

    if country_code is not None:
        custom_day.country_code = (country_code or '').strip().upper() or custom_day.country_code

    if is_recurring is not None:
        custom_day.is_recurring = bool(is_recurring)

    custom_day.save()
    return JsonResponse({'customDay': _serialize_custom_special_day(custom_day)}, encoder=DjangoJSONEncoder)


def _get_special_day_events(year, month):
    cache_key = f"special_day_events_{year}_{month}"
    cached = cache.get(cache_key)
    if cached is not None:
        return copy.deepcopy(cached)

    events = []
    DateTime = datetime_module.datetime

    accurate_nigerian_holidays = get_accurate_nigerian_holidays(year)
    for holiday in accurate_nigerian_holidays:
        try:
            holiday_date = DateTime.strptime(holiday['date'], '%Y-%m-%d')
            if holiday_date.month == month:
                events.append({
                    'title': f"{holiday['name']} (NG)",
                    'start': holiday['date'],
                    'allDay': True,
                    'extendedProps': {
                        'source': 'ng',
                        'name': holiday['name'],
                        'countryCode': 'NG',
                        'accurate': True,
                        'type': 'national',
                        'category': 'national',
                    },
                    'classNames': ['holiday', 'nigeria'],
                })
        except Exception:
            continue

    try:
        r = requests.get(f'https://date.nager.at/api/v3/PublicHolidays/{year}/NG', timeout=6)
        if r.ok:
            ng_list = r.json()
            for h in ng_list:
                try:
                    d = h.get('date')
                    dt = DateTime.fromisoformat(d) if 'T' in d else DateTime.strptime(d, '%Y-%m-%d')
                    if dt.month != month:
                        continue

                    holiday_name = h.get('localName') or h.get('name')
                    if not any(
                        e['extendedProps'].get('name') == holiday_name
                        and e['extendedProps'].get('accurate')
                        for e in events
                    ):
                        events.append({
                            'title': f"{holiday_name} (NG)",
                            'start': d,
                            'allDay': True,
                            'extendedProps': {
                                'source': 'ng',
                                'name': holiday_name,
                                'countryCode': 'NG',
                                'api_source': True,
                                'type': 'national',
                                'category': 'national',
                            },
                            'classNames': ['holiday', 'nigeria'],
                        })
                except Exception as exc:
                    print(f"Error processing Nigerian holiday: {exc}")
                    continue
    except Exception as exc:
        print(f"Nigerian holidays API failed: {exc}")

    try:
        r2 = requests.get('https://date.nager.at/api/v3/NextPublicHolidaysWorldwide', timeout=6)
        if r2.ok:
            world = r2.json()
            for h in world:
                try:
                    d = h.get('date')
                    dt = DateTime.fromisoformat(d) if 'T' in d else DateTime.strptime(d, '%Y-%m-%d')
                    if dt.year != year or dt.month != month:
                        continue
                    if h.get('countryCode') == 'NG':
                        continue

                    events.append({
                        'title': f"{h.get('localName') or h.get('name')} ({h.get('countryCode')})",
                        'start': d,
                        'allDay': True,
                        'extendedProps': {
                            'source': 'global',
                            'name': h.get('localName') or h.get('name'),
                            'countryCode': h.get('countryCode'),
                        },
                        'classNames': ['holiday', 'global'],
                    })
                except Exception as exc:
                    print(f"Error processing global holiday: {exc}")
                    continue
    except Exception as exc:
        print(f"Global holidays API failed: {exc}")

    try:
        from .models import Holiday, CustomSpecialDay
        from django.db.models import Q

        db_holidays = Holiday.objects.filter(
            Q(recurring=True, date__month=month) |
            Q(recurring=False, date__year=year, date__month=month)
        )

        def _normalize_recurring_date(source_date):
            try:
                return date(year, source_date.month, source_date.day)
            except ValueError:
                if source_date.month == 2 and source_date.day == 29:
                    # Use Feb 28 for non-leap years to retain visibility of leap events
                    return date(year, 2, 28)
                raise

        for holiday in db_holidays:
            try:
                event_date = holiday.date if not holiday.recurring else _normalize_recurring_date(holiday.date)
            except Exception as exc:
                print(f"Failed to normalize holiday '{holiday.name}': {exc}")
                continue

            events.append({
                'id': f"local-{holiday.id}",
                'title': f"{holiday.name} (Local)",
                'start': event_date.isoformat(),
                'allDay': True,
                'extendedProps': {
                    'source': 'local',
                    'name': holiday.name,
                    'countryCode': 'NG',
                    'local_db': True,
                    'type': (holiday.category or 'custom'),
                    'recurring': holiday.recurring,
                    'description': holiday.description or '',
                    'category': (holiday.category or 'custom'),
                    'eventId': holiday.id,
                },
                'classNames': ['holiday', 'local'],
            })
    except Exception as exc:
        print(f"Local holiday database query failed: {exc}")

    try:
        from django.db.models import Q

        custom_days = CustomSpecialDay.objects.filter(
            Q(is_recurring=True, date__month=month) |
            Q(is_recurring=False, date__year=year, date__month=month)
        )

        for custom in custom_days:
            try:
                if custom.is_recurring:
                    event_date = date(year, custom.date.month, custom.date.day)
                else:
                    event_date = custom.date
            except Exception as exc:
                print(f"Failed to normalize custom special day '{custom.name}': {exc}")
                continue

            events.append({
                'id': f"custom-{custom.id}",
                'title': custom.name,
                'start': event_date.isoformat(),
                'allDay': True,
                'extendedProps': {
                    'source': 'custom',
                    'name': custom.name,
                    'countryCode': (custom.country_code or '').upper() or 'NG',
                    'local_db': True,
                    'type': custom.category or 'custom',
                    'category': custom.category or 'custom',
                    'description': custom.description or '',
                    'recurring': custom.is_recurring,
                    'eventId': str(custom.id),
                    'custom_admin': True,
                },
                'classNames': ['holiday', 'custom'],
            })
    except Exception as exc:
        print(f"Custom special day query failed: {exc}")

    def _is_nigeria_related(event):
        props = (event.get('extendedProps') or {})
        country = (props.get('countryCode') or '').upper()
        if country == 'NG':
            return True
        source = (props.get('source') or '').lower()
        if source in {'ng', 'nigeria', 'nigerian', 'local'}:
            return True
        if props.get('local_db'):
            return True
        if props.get('custom_admin'):
            return True
        return False

    events = [e for e in events if _is_nigeria_related(e)]

    seen = set()
    uniq = []
    for e in events:
        key = f"{e['start']}::{e['extendedProps'].get('name', '').lower()}"
        if key not in seen:
            seen.add(key)
            uniq.append(e)

    uniq.sort(key=lambda x: x['start'])
    cache.set(cache_key, uniq, 6 * 60 * 60)  # cache for 6 hours
    return copy.deepcopy(uniq)


@require_http_methods(['GET'])
@login_required
@support_required
def api_special_days(request):
    """Return calendar events for special days."""
    try:
        today_lagos = lagos_today()
        year = int(request.GET.get('year', today_lagos.year))
        month = int(request.GET.get('month', today_lagos.month))
    except Exception:
        today_lagos = lagos_today()
        year = today_lagos.year
        month = today_lagos.month

    events = _get_special_day_events(year, month)
    return JsonResponse({'events': events}, encoder=DjangoJSONEncoder)


def _collect_special_days_by_date(events):
    grouped = {}
    for event in events:
        start = (event.get('start') or '').split('T')[0]
        if not start:
            continue
        grouped.setdefault(start, []).append(event)
    return grouped


@require_http_methods(['GET'])
def api_special_days_summary(request):
    support_user = _authenticate_support_request(request)
    if not support_user:
        return JsonResponse({'detail': 'Authentication credentials were not provided.'}, status=401)

    today = timezone.localdate()
    week_end = today + timedelta(days=6)

    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except Exception:
        year, month = today.year, today.month

    events = _get_special_day_events(year, month)
    dedup_keys = {
        f"{e['start']}::{e['extendedProps'].get('name', '').lower()}"
        for e in events
    }
    if week_end.year != year or week_end.month != month:
        extra_events = _get_special_day_events(week_end.year, week_end.month)
        for event in extra_events:
            key = f"{event['start']}::{event['extendedProps'].get('name', '').lower()}"
            if key not in dedup_keys:
                events.append(event)
                dedup_keys.add(key)
    grouped = _collect_special_days_by_date(events)

    def within_range(start_date, end_date):
        window = []
        current = start_date
        while current <= end_date:
            window.append(current.isoformat())
            current += timedelta(days=1)
        return window

    today_key = today.isoformat()
    week_keys = within_range(today, week_end)

    def normalize(event):
        props = event.get('extendedProps', {})
        title = event.get('title') or props.get('name') or 'Event'
        source = (props.get('source') or '').lower()
        is_local = bool(props.get('local_db'))
        is_nigerian = props.get('countryCode') == 'NG' or 'nigerian' in source or source == 'ng' or is_local
        event_type = (props.get('type') or ('custom' if is_local else ('national' if is_nigerian else 'custom'))).lower()
        category = (props.get('category') or event_type or 'custom').lower()
        include = is_nigerian or event_type in {'national', 'custom', 'commercial', 'observance', 'religious'} or is_local
        if not include:
            return None
        description = props.get('description') or event.get('description') or title
        return {
            'id': event.get('id'),
            'name': props.get('name') or title,
            'title': title,
            'date': (event.get('start') or '').split('T')[0],
            'description': description,
            'countryCode': props.get('countryCode') or ('NG' if is_nigerian else ''),
            'type': event_type,
            'category': category,
            'source': props.get('source'),
            'isNigerian': is_nigerian,
            'isCustom': is_local or category == 'custom',
            'isRecurring': bool(props.get('recurring')),
            'localEventId': props.get('eventId'),
        }

    today_events = []
    for ev in grouped.get(today_key, []):
        normalized = normalize(ev)
        if normalized is not None:
            today_events.append(normalized)

    week_events = []
    for key in week_keys:
        for ev in grouped.get(key, []):
            normalized = normalize(ev)
            if normalized is not None:
                week_events.append(normalized)

    month_events = []
    month_seen = set()
    month_prefix = f"{year:04d}-{month:02d}-"
    month_keys = sorted(k for k in grouped.keys() if k.startswith(month_prefix))
    for key in month_keys:
        for ev in grouped.get(key, []):
            normalized = normalize(ev)
            if normalized is None:
                continue
            dedup_key = (
                normalized.get('date'),
                (normalized.get('name') or normalized.get('title') or '').lower(),
            )
            if dedup_key in month_seen:
                continue
            month_seen.add(dedup_key)
            month_events.append(normalized)

    def _sort_key(item):
        return (item.get('date') or '', item.get('name') or '')

    today_events.sort(key=_sort_key)
    week_events.sort(key=_sort_key)
    month_events.sort(key=_sort_key)

    response = {
        'generatedAt': timezone.now().isoformat(),
        'today': today_events,
        'thisWeek': week_events,
        'thisMonth': month_events,
        'weekRange': {
            'start': today.isoformat(),
            'end': week_end.isoformat(),
        },
        'month': month,
        'year': year,
    }

    return JsonResponse(response, encoder=DjangoJSONEncoder)


def get_accurate_nigerian_holidays(year):
    """
    Returns accurate Nigerian holidays for the given year
    These are the official dates that don't depend on external APIs
    """
    # Base holidays that occur on the same date every year
    fixed_holidays = [
        {'date': f'{year}-01-01', 'name': "New Year's Day"},
        {'date': f'{year}-05-01', 'name': "Workers' Day"},
        {'date': f'{year}-06-12', 'name': "Democracy Day"},
        {'date': f'{year}-10-01', 'name': "Independence Day"},  # Correct: October 1st
        {'date': f'{year}-12-25', 'name': "Christmas Day"},
        {'date': f'{year}-12-26', 'name': "Boxing Day"},
    ]
    
    # Year-specific holidays (these dates were researched for accuracy)
    year_specific = {
        2025: [
            {'date': f'{year}-04-18', 'name': "Good Friday"},
            {'date': f'{year}-04-21', 'name': "Easter Monday"},
            {'date': f'{year}-03-31', 'name': "Eid al-Fitr"}, # Estimated
            {'date': f'{year}-06-07', 'name': "Eid al-Adha"}, # Estimated
        ],
        2024: [
            {'date': f'{year}-03-29', 'name': "Good Friday"},
            {'date': f'{year}-04-01', 'name': "Easter Monday"},
            {'date': f'{year}-04-10', 'name': "Eid al-Fitr"},
            {'date': f'{year}-06-16', 'name': "Eid al-Adha"},
        ]
    }
    
    holidays = fixed_holidays.copy()
    if year in year_specific:
        holidays.extend(year_specific[year])
    
    return holidays


# SEND MESSAGE
@require_http_methods(['GET', 'PATCH', 'DELETE'])
@login_required
@support_required
def api_templates_detail(request, tpl_id):
    """
    REST detail for MessageTemplate (this is what the front-end expects at /api/templates/<uuid>/)
    GET -> return template
    PATCH -> partial update (name, channel, audience, subject, body, send_time, is_active)
    DELETE -> delete template
    """
    tpl = get_object_or_404(MessageTemplate, id=tpl_id)

    if request.method == "GET":
        data = {
            'id': tpl.id,
            'name': tpl.name,
            'channel': tpl.channel,
            'audience': tpl.audience,
            'subject': tpl.subject,
            'message': tpl.body,
            'time': tpl.send_time.isoformat() if tpl.send_time else None,
            'isActive': tpl.is_active,
            'createdAt': tpl.created_at.isoformat() if tpl.created_at else None,
        }
        return JsonResponse({'template': data}, encoder=DjangoJSONEncoder)

    if request.method == "DELETE":
        tpl.delete()
        ActivityLog.objects.create(actor=request.user, action=f"Deleted template '{tpl.name}'")
        return HttpResponse(status=204)

    # PATCH
    if request.method == "PATCH":
        try:
            payload = json.loads(request.body.decode('utf-8')) if request.body else {}
        except Exception:
            return HttpResponseBadRequest("Invalid JSON")

        if 'name' in payload:
            tpl.name = payload.get('name') or tpl.name
        if 'channel' in payload:
            tpl.channel = payload.get('channel') or tpl.channel
        if 'audience' in payload:
            tpl.audience = payload.get('audience') or tpl.audience
        if 'subject' in payload:
            tpl.subject = payload.get('subject') or tpl.subject
        if 'message' in payload:
            tpl.body = payload.get('message') or tpl.body
        if 'isActive' in payload:
            tpl.is_active = bool(payload.get('isActive'))
        if 'time' in payload:
            t = payload.get('time')
            if t:
                try:
                    from datetime import time as dt_time
                    parts = str(t).split(':')
                    if len(parts) >= 2:
                        hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                        tpl.send_time = dt_time(hour, minute, second)
                except Exception:
                    pass
            else:
                tpl.send_time = None
        tpl.save()
        ActivityLog.objects.create(actor=request.user, action=f"Updated template '{tpl.name}'")
        return JsonResponse({'ok': True, 'id': tpl.id}, encoder=DjangoJSONEncoder)

@require_http_methods(['GET', 'PATCH', 'DELETE'])
@login_required
@support_required
def api_messages_detail(request, msg_id):
    """
    GET -> return single InAppMessage (with minimal metadata)
    PATCH -> allow updating subject/body/is_draft (sender must be current user for safety)
    DELETE -> delete (sender must be current user)
    """
    msg = get_object_or_404(InAppMessage, id=msg_id)

    # only sender may edit/delete drafts/messages
    if request.method in ('PATCH', 'DELETE') and msg.sender_id != request.user.id:
        return HttpResponse(status=403)

    if request.method == "GET":
        data = {
            'id': msg.id,
            'subject': msg.subject,
            'body': msg.body,
            'isDraft': bool(msg.is_draft),
            'createdAt': msg.created_at.isoformat() if msg.created_at else None,
            'recipients': list(msg.recipients.values_list('id', flat=True))[:200],
        }
        return JsonResponse({'message': data}, encoder=DjangoJSONEncoder)

    if request.method == "DELETE":
        msg.delete()
        ActivityLog.objects.create(actor=request.user, action=f"Deleted InApp message id={msg_id}")
        return JsonResponse({'ok': True})

    # PATCH
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    if 'subject' in payload:
        msg.subject = payload.get('subject') or msg.subject
    if 'body' in payload:
        msg.body = payload.get('body') or msg.body
    if 'is_draft' in payload or 'isDraft' in payload:
        msg.is_draft = bool(payload.get('is_draft', payload.get('isDraft', msg.is_draft)))
    msg.save()
    ActivityLog.objects.create(actor=request.user, action=f"Updated InApp message id={msg.id}")
    return JsonResponse({'ok': True, 'id': msg.id}, encoder=DjangoJSONEncoder)

@require_http_methods(['GET', 'PATCH', 'DELETE'])
@login_required
@support_required
def api_outbound_detail(request, outbound_id):
    """
    GET -> return outbound with small items preview
    PATCH -> update body/subject/status
    DELETE -> delete queued outbound
    """
    om = get_object_or_404(OutboundMessage, id=outbound_id)

    if request.method == "GET":
        data = {
            'id': str(om.id),
            'channel': om.channel,
            'audience': om.audience,
            'subject': om.subject,
            'body': om.body,
            'status': om.status,
            'createdAt': om.created_at.isoformat() if om.created_at else None,
            'scheduledFor': om.scheduled_for.isoformat() if om.scheduled_for else None,
            'toCount': om.recipients_count,
            'toSample': om.recipients_sample or [],
            'itemsPreview': []  # optionally return items
        }
        # include up to 20 items previews if you want
        try:
            items = om.items.all()[:20]
            data['itemsPreview'] = [{'id': it.id, 'recipient_id': it.recipient_id, 'status': it.status} for it in items]
        except Exception:
            pass
        return JsonResponse({'outbound': data}, encoder=DjangoJSONEncoder)

    if request.method == "DELETE":
        om.delete()
        ActivityLog.objects.create(actor=request.user, action=f"Deleted outbound id={outbound_id}")
        return JsonResponse({'ok': True})

    # PATCH
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    changed = False
    if 'body' in payload:
        om.body = payload.get('body') or om.body; changed = True
    if 'subject' in payload:
        om.subject = payload.get('subject') or om.subject; changed = True
    if 'status' in payload:
        om.status = payload.get('status') or om.status; changed = True
    if changed:
        om.save()
        ActivityLog.objects.create(actor=request.user, action=f"Updated outbound id={om.id}")
    return JsonResponse({'ok': True, 'id': str(om.id)}, encoder=DjangoJSONEncoder)

# NEWSLETTER

@require_http_methods(['POST'])
@login_required
@support_required
def api_newsletter_preview(request):
    """
    POST JSON: { subject, audience, html_body }
    Returns sanitized HTML preview (server-side) + a small recipients sample.
    """
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    subject = payload.get('subject') or ''
    audience = payload.get('audience') or 'all'
    html_body = payload.get('html_body') or payload.get('html') or payload.get('body') or ''

    # Sanitize HTML (best-effort). If bleach available, use it.
    if BLEACH_AVAILABLE:
        # Expand allowed tags a bit for email templates
        allowed_tags = list(bleach.sanitizer.ALLOWED_TAGS) + ['p', 'br', 'img', 'h1', 'h2', 'h3', 'h4',
                                                             'table', 'thead', 'tbody', 'tr', 'td', 'th', 'caption']
        allowed_attrs = {'a': ['href', 'title', 'target', 'rel'], 'img': ['src', 'alt', 'width', 'height'], '*': ['class', 'style']}
        try:
            clean_html = bleach.clean(html_body, tags=allowed_tags, attributes=allowed_attrs, strip=True)
        except Exception:
            clean_html = html_body
    else:
        # No bleach installed — return body but mark as not sanitized on client if you need to flag it
        clean_html = html_body

    # Small recipients sample (max 6)
    if audience == 'clients':
        qs = User.objects.filter(role='client').exclude(email__isnull=True).exclude(email__exact='')[:6]
    elif audience == 'marketers':
        qs = User.objects.filter(role='marketer').exclude(email__isnull=True).exclude(email__exact='')[:6]
    else:
        qs = User.objects.exclude(email__isnull=True).exclude(email__exact='')[:6]

    sample = [{'id': u.id, 'name': (u.get_full_name() or u.email), 'email': u.email} for u in qs]

    return JsonResponse({'subject': subject, 'html': clean_html, 'sample': sample}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_newsletter_send(request):
    """
    POST JSON payload:
    {
      "subject": "...",
      "audience": "clients|marketers|all",
      "html_body": "<h1>..</h1>",
      "send_now": true|false,
      "scheduled_for": "ISO8601 datetime string (optional)",
      "name": "Campaign name (optional)"
    }

    Creates NewsletterCampaign and OutboundMessage. Optionally dispatches a send task.
    """
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    subject = (payload.get('subject') or '').strip()
    audience = payload.get('audience') or 'all'
    html_body = payload.get('html_body') or payload.get('html') or ''
    send_now = bool(payload.get('send_now', True))
    scheduled_for = payload.get('scheduled_for')
    name = payload.get('name') or subject[:120]

    if not subject or not html_body:
        return JsonResponse({'ok': False, 'error': 'subject and html_body required'}, status=400)

    # Create campaign (initial status: queued if send_now else draft)
    campaign = NewsletterCampaign.objects.create(
        name=name,
        subject=subject,
        html_body=html_body,
        audience=audience,
        created_by=request.user,
        status=NewsletterCampaign.STATUS_QUEUED if send_now else NewsletterCampaign.STATUS_DRAFT
    )

    # Parse scheduled_for if present (expect ISO8601)
    if scheduled_for:
        try:
            parsed = datetime.fromisoformat(scheduled_for)
            if timezone.is_naive(parsed):
                parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
            campaign.scheduled_for = parsed
            campaign.save(update_fields=['scheduled_for'])
        except Exception:
            # Ignore parse error
            pass

    # Compute recipients and sample (light summary)
    if audience == 'clients':
        recipients_qs = User.objects.filter(role='client').exclude(email__isnull=True).exclude(email__exact='')
    elif audience == 'marketers':
        recipients_qs = User.objects.filter(role='marketer').exclude(email__isnull=True).exclude(email__exact='')
    else:
        recipients_qs = User.objects.exclude(email__isnull=True).exclude(email__exact='')

    recipients_count = recipients_qs.count()
    sample = list(recipients_qs.values_list('first_name', 'last_name', 'email')[:5])
    sample_list = [f"{(a or '').strip()} {(b or '').strip()}".strip() + (f" <{c}>" if c else '') for a, b, c in sample]

    campaign.recipients_count = recipients_count
    campaign.recipients_sample = sample_list
    campaign.save(update_fields=['recipients_count', 'recipients_sample'])

    # Create OutboundMessage row for visibility/queue (worker can expand outbound->items)
    om = OutboundMessage.objects.create(
        channel='email',
        audience=audience,
        subject=subject,
        body=html_body,
        created_by=request.user,
        recipients_count=recipients_count,
        recipients_sample=sample_list,
        status='queued',
        scheduled_for=campaign.scheduled_for
    )

    dispatched = False
    # Attempt to dispatch background task if available
    try:
        from .tasks import send_newsletter_task  # <-- optional celery task or function you implement
        if send_now:
            # If celery task object with .delay() is present -> use .delay(); otherwise call synchronously
            if hasattr(send_newsletter_task, 'delay'):
                send_newsletter_task.delay(str(campaign.id))
            else:
                # Synchronous fallback executed in request (not recommended for large lists)
                send_newsletter_task(str(campaign.id))
            dispatched = True
    except Exception as e:
        # Record error but still return success (campaign and outbound created)
        campaign.error = str(e)[:2000]
        campaign.save(update_fields=['error'])

    return JsonResponse({'ok': True, 'id': str(campaign.id), 'outbound_id': str(om.id), 'dispatched': dispatched}, encoder=DjangoJSONEncoder)


@require_POST
@login_required
@support_required
def api_upload_image(request):
    """
    Handles image uploads from editor (TinyMCE). Returns JSON:
    { "location": "/media/uploads/..." }  (TinyMCE expects `location`)
    """
    upload = request.FILES.get('file') or request.FILES.get('image') or None
    if not upload:
        return HttpResponseBadRequest("No file uploaded")

    # Basic content type check
    ALLOWED = {'image/jpeg', 'image/png', 'image/gif', 'image/webp'}
    if upload.content_type not in ALLOWED:
        return HttpResponseBadRequest("Unsupported file type")

    # Create a safe filename
    fn = get_valid_filename(upload.name)
    base, ext = os.path.splitext(fn)
    filename = f"uploads/newsletter/{uuid.uuid4().hex}{ext.lower()}"

    # Use default_storage (local MEDIA or S3 if configured)
    saved = default_storage.save(filename, ContentFile(upload.read()))
    url = default_storage.url(saved)

    # Return location as TinyMCE expects
    return JsonResponse({"location": url})

# SETTINGS
@login_required
@support_required
def settings(request):
    ctx = {'home_url': '/'}
    return render(request, 'adminSupport/customer_relation/settings.html', ctx)


@require_http_methods(['GET'])
@login_required
@support_required
def api_settings_get(request):
    """
    Return current settings + lists of birthday/special templates for the UI pickers.
    """
    ms, _ = MessagingSettings.objects.get_or_create(pk=1)
    # available templates
    b_templates = list(BirthdayTemplate.objects.filter(template_type=BirthdayTemplate.TEMPLATE_TYPE_BIRTHDAY)
                       .order_by('-created_at').values('id', 'name', 'channel', 'audience'))
    s_templates = list(BirthdayTemplate.objects.filter(template_type=BirthdayTemplate.TEMPLATE_TYPE_SPECIAL)
                       .order_by('-created_at').values('id', 'name', 'channel', 'audience'))

    # serialize minimal
    data = {
        'messaging': {
            'auto_birthday': bool(ms.auto_birthday),
            'auto_special_day': bool(ms.auto_special_day),
            'updated_at': ms.updated_at.isoformat() if ms.updated_at else None,
        },
        'birthday_templates': b_templates,
        'special_templates': s_templates,
    }
    # If you extended MessagingSettings with fields like birthday_template_id/time/channel, include them here.
    # e.g. data['messaging']['birthday_template_id'] = str(ms.birthday_template.id) if ms.birthday_template else None

    return JsonResponse(data, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_settings_update(request):
    """
    Accepts JSON payload:
    {
      "messaging": {
         "auto_birthday": true,
         "auto_special_day": false,
         "birthday_template_id": "<uuid or null>",
         "birthday_time": "HH:MM" (optional),
         "birthday_channel": "sms|email|inapp",
         "special_template_id": "...",
         "special_time": "HH:MM",
         "special_channel": "..."
      },
      "newsletter": {
         "batch_size": 100,
         "use_celery": true
      }
    }
    """
    try:
        payload = json.loads(request.body.decode('utf-8') or '{}')
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    ms, _ = MessagingSettings.objects.get_or_create(pk=1)
    messaging = payload.get('messaging', {})

    changed = False
    # toggles
    if 'auto_birthday' in messaging:
        ms.auto_birthday = bool(messaging.get('auto_birthday'))
        changed = True
    if 'auto_special_day' in messaging:
        ms.auto_special_day = bool(messaging.get('auto_special_day'))
        changed = True

    # optional: if you added birthday_template in MessagingSettings
    if 'birthday_template_id' in messaging:
        tpl_id = messaging.get('birthday_template_id') or None
        if tpl_id:
            try:
                ms.birthday_template = BirthdayTemplate.objects.get(id=tpl_id)
            except BirthdayTemplate.DoesNotExist:
                ms.birthday_template = None
        else:
            ms.birthday_template = None
        changed = True

    if 'special_template_id' in messaging:
        tpl_id = messaging.get('special_template_id') or None
        if tpl_id:
            try:
                ms.special_template = BirthdayTemplate.objects.get(id=tpl_id)
            except BirthdayTemplate.DoesNotExist:
                ms.special_template = None
        else:
            ms.special_template = None
        changed = True

    # optional time/channel handling (if fields exist)
    if 'birthday_time' in messaging:
        t = messaging.get('birthday_time')
        if t:
            try:
                from datetime import time as dt_time
                parts = t.split(':')
                hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                ms.birthday_time = dt_time(hour, minute, second)
            except Exception:
                pass
        else:
            ms.birthday_time = None
        changed = True

    if 'special_time' in messaging:
        t = messaging.get('special_time')
        if t:
            try:
                from datetime import time as dt_time
                parts = t.split(':')
                hour = int(parts[0]); minute = int(parts[1]); second = int(parts[2]) if len(parts) > 2 else 0
                ms.special_time = dt_time(hour, minute, second)
            except Exception:
                pass
        else:
            ms.special_time = None
        changed = True

    # Save if changed
    if changed:
        ms.save()
        ActivityLog.objects.create(actor=request.user, action=f"Updated messaging settings by {request.user.get_full_name() or request.user.username}")

    # newsletter section (not stored in current models — you can store as site config or environment)
    newsletter = payload.get('newsletter', {})
    # e.g. if you add a NewsletterSettings model, save here

    return JsonResponse({'ok': True}, encoder=DjangoJSONEncoder)


@require_http_methods(['POST'])
@login_required
@support_required
def api_settings_run_triggers(request):
    """
    Run scheduled triggers immediately (wraps your existing run_send_birthday_messages / api_run_triggers).
    """
    # if you already have api_run_triggers, you can call it or reuse its logic.
    # Here we call the existing helper if available.
    try:
        from .tasks import run_send_birthday_messages
        result = run_send_birthday_messages()
        ActivityLog.objects.create(actor=request.user, action=f"Manually ran scheduled triggers (by {request.user.username})")
        return JsonResponse({'ok': True, 'result': result}, encoder=DjangoJSONEncoder)
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)


# APP CONTENT MANAGEMENT
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from django.shortcuts import get_object_or_404
from .models import (
    SiteTheme, Statistic, Project, Update, 
    DiasporaOffer, SocialProof, ContactInfo, 
    WheelPrize, SpinLimit, ThemeRule, Holiday, Season
)
from django.utils import timezone
from datetime import date
import json

def get_active_theme():
    """Determine the active theme based on current date and rules"""
    today = date.today()
    
    # Get all active themes that are applicable today
    applicable_themes = []
    for theme in SiteTheme.objects.filter(is_active=True):
        if theme.is_applicable_today():
            applicable_themes.append(theme)
    
    # Also check theme rules
    for rule in ThemeRule.objects.filter(is_active=True):
        if rule.evaluate(today):
            applicable_themes.append(rule.theme)
    
    # Sort by priority (highest first) and return the first one
    if applicable_themes:
        applicable_themes.sort(key=lambda x: x.priority, reverse=True)
        return applicable_themes[0]
    
    # Fallback to default theme
    return SiteTheme.objects.filter(theme_type='default', is_active=True).first()

@method_decorator(csrf_exempt, name='dispatch')
class LandingPageData(View):
    def get(self, request):
        # Get active theme
        active_theme = get_active_theme()
        
        # If no theme found, create a default one
        if not active_theme:
            active_theme = SiteTheme.objects.create(
                name="Default Theme",
                theme_type="default",
                primary_color="#6A5AE0",
                accent_color="#FFFFFF",
                hero_image="https://images.unsplash.com/photo-1494526585095-c41746248156?w=800&idx=i",
                banner_text="Welcome to Lior & Eliora Properties",
                gradient_colors=["#00000040", "#6A5AE014"],
                gradient_stops=[0.0, 1.0],
                gradient_begin="topCenter",
                gradient_end="bottomCenter",
                accent_gradient_colors=["#00FFFF", "#008080"],
                is_active=True
            )
        
        # Get other data (same as before)
        stats = Statistic.objects.first() or Statistic.objects.create()
        featured_project = Project.objects.filter(is_featured=True).first()
        updates = Update.objects.filter(is_active=True)[:5]
        diaspora_offers = DiasporaOffer.objects.filter(is_active=True).order_by('order')
        social_proofs = SocialProof.objects.filter(is_active=True).order_by('order')[:4]
        contact_info = ContactInfo.objects.first() or ContactInfo.objects.create(
            company_name="Lior & Eliora Properties",
            cac_number="RC123456789",
            address="12 Victoria Island, Lagos",
            whatsapp_number="+2348012345678",
            phone_number="+2348012345678"
        )
        wheel_prizes = WheelPrize.objects.filter(is_active=True).order_by('order')
        spin_limit = SpinLimit.objects.first() or SpinLimit.objects.create()
        
        # Build response data
        data = {
            'theme': {
                'name': active_theme.name,
                'type': active_theme.theme_type,
                'primary': active_theme.primary_color,
                'secondary': active_theme.secondary_color or active_theme.primary_color,
                'accent': active_theme.accent_color,
                'hero': active_theme.hero_image,
                'bannerText': active_theme.banner_text,
                'gradient': {
                    'colors': active_theme.gradient_colors,
                    'stops': active_theme.gradient_stops,
                    'begin': active_theme.gradient_begin,
                    'end': active_theme.gradient_end,
                },
                'accentGradient': {
                    'colors': active_theme.accent_gradient_colors,
                }
            },
            # Other data remains the same as before
            'stats': {
                'homesDelivered': stats.homes_delivered,
                'estatesDeveloped': stats.estates_developed,
                'yearsInBusiness': stats.years_in_business,
            },
            'project': {
                'name': featured_project.name if featured_project else 'Victoria Court Phase 2',
                'location': featured_project.location if featured_project else 'Ajah',
                'soldPercent': featured_project.sold_percent if featured_project else 70,
                'image': featured_project.image if featured_project else 'https://images.unsplash.com/photo-1568605114967-8130f3a36994?w=1200',
            },
            'updates': [
                {
                    'title': update.title,
                    'date': update.date.strftime('%b %d'),
                    'image': update.image,
                    'description': update.description,
                }
                for update in updates
            ],
            'diasporaOffers': [
                {
                    'title': offer.title,
                    'icon': offer.icon,
                    'subtitle': offer.subtitle,
                    'detail': offer.detail,
                    'isVirtual': offer.is_virtual,
                }
                for offer in diaspora_offers
            ],
            'socialProofs': [
                {
                    'image': proof.image,
                    'caption': proof.caption,
                }
                for proof in social_proofs
            ],
            'contact': {
                'companyName': contact_info.company_name,
                'cacNumber': contact_info.cac_number,
                'address': contact_info.address,
                'whatsappNumber': contact_info.whatsapp_number,
                'phoneNumber': contact_info.phone_number,
                'facebook': contact_info.facebook_url,
                'instagram': contact_info.instagram_url,
                'linkedin': contact_info.linkedin_url,
                'youtube': contact_info.youtube_url,
            },
            'wheelPrizes': [prize.text for prize in wheel_prizes],
            'spinLimit': spin_limit.spins_per_day,
        }
        
        return JsonResponse(data)

# Additional API endpoint to get all themes
@method_decorator(csrf_exempt, name='dispatch')
class ThemeList(View):
    def get(self, request):
        themes = SiteTheme.objects.filter(is_active=True)
        
        theme_data = []
        for theme in themes:
            theme_data.append({
                'id': theme.id,
                'name': theme.name,
                'type': theme.theme_type,
                'primary_color': theme.primary_color,
                'secondary_color': theme.secondary_color,
                'accent_color': theme.accent_color,
                'hero_image': theme.hero_image,
                'banner_text': theme.banner_text,
                'is_applicable_today': theme.is_applicable_today(),
                'priority': theme.priority,
            })
            
        return JsonResponse({'themes': theme_data})


@method_decorator(csrf_exempt, name='dispatch')
class SpinWheel(View):
    def post(self, request):
        # This would typically require user authentication
        # For now, we'll just return a random prize
        import random
        
        wheel_prizes = WheelPrize.objects.filter(is_active=True)
        if not wheel_prizes:
            return JsonResponse({'error': 'No prizes available'}, status=400)
        
        # Select a random prize
        prize = random.choice(wheel_prizes)
        
        return JsonResponse({
            'prize': prize.text,
            'message': 'Congratulations!'
        })


# ===================== ENTERPRISE CRM VIEWS =====================

@login_required
def crm_enterprise_dashboard(request):
    """Enterprise CRM Dashboard with advanced features"""
    context = {
        'page_title': 'Enterprise CRM Dashboard',
        'breadcrumb': [
            {'name': 'Admin Support', 'url': '/admin-support/'},
            {'name': 'Customer Relations', 'url': '/admin-support/support/'},
            {'name': 'Enterprise Dashboard', 'url': ''},
        ]
    }
    return render(request, 'adminSupport/customer_relation/crm_enterprise_dashboard.html', context)


@login_required
def special_days_manager(request):
    """Special Days Manager for holiday and event campaigns"""
    context = {
        'page_title': 'Special Days Manager',
        'breadcrumb': [
            {'name': 'Admin Support', 'url': '/admin-support/'},
            {'name': 'Customer Relations', 'url': '/admin-support/support/'},
            {'name': 'Special Days Manager', 'url': ''},
        ]
    }
    return render(request, 'adminSupport/customer_relation/special_days_manager.html', context)


@login_required
def campaign_builder(request):
    """Advanced Campaign Builder"""
    context = {
        'page_title': 'Campaign Builder',
        'breadcrumb': [
            {'name': 'Admin Support', 'url': '/admin-support/'},
            {'name': 'Customer Relations', 'url': '/admin-support/support/'},
            {'name': 'Campaign Builder', 'url': ''},
        ]
    }
    return render(request, 'adminSupport/customer_relation/campaign_builder.html', context)


@login_required
def audience_segments(request):
    """Audience Segmentation Manager"""
    context = {
        'page_title': 'Audience Segments',
        'breadcrumb': [
            {'name': 'Admin Support', 'url': '/admin-support/'},
            {'name': 'Customer Relations', 'url': '/admin-support/support/'},
            {'name': 'Audience Segments', 'url': ''},
        ]
    }
    return render(request, 'adminSupport/customer_relation/audience_segments.html', context)


@login_required
def crm_analytics(request):
    """Advanced CRM Analytics Dashboard"""
    context = {
        'page_title': 'CRM Analytics',
        'breadcrumb': [
            {'name': 'Admin Support', 'url': '/admin-support/'},
            {'name': 'Customer Relations', 'url': '/admin-support/support/'},
            {'name': 'Analytics', 'url': ''},
        ]
    }
    return render(request, 'adminSupport/customer_relation/crm_analytics.html', context)


@login_required
def automation_settings(request):
    """CRM Automation Settings"""
    context = {
        'page_title': 'Automation Settings',
        'breadcrumb': [
            {'name': 'Admin Support', 'url': '/admin-support/'},
            {'name': 'Customer Relations', 'url': '/admin-support/support/'},
            {'name': 'Automation Settings', 'url': ''},
        ]
    }
    return render(request, 'adminSupport/customer_relation/automation_settings.html', context)
